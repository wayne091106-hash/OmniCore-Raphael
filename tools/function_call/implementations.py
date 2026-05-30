"""
工具實作庫 — 從 nim_tool_test/tools/implementations.py 直搬
修改：
  1. import 路徑：tools.registry → .registry（相對 import）
  2. Google 憑證路徑：parent.parent → parent.parent.parent（指向專案根目錄）
"""
import os
import math
import atexit
import base64
import contextlib
import ctypes
import datetime
import csv
import hashlib
import json
import mimetypes
import platform
import socket
import sqlite3
import urllib.request
import urllib.error
import urllib.parse
import subprocess
import shutil
import winreg
import re
import sys
import time
import uuid
import zipfile
from pathlib import Path
from ctypes import wintypes
from dotenv import load_dotenv
from .registry import registry, tool

load_dotenv()

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DATA_DIR = PROJECT_ROOT / "data"
UPLOAD_DIR = DATA_DIR / "uploads"
OUTPUT_DIR = DATA_DIR / "outputs"
SITE_MEMORY_FILE = DATA_DIR / "site_memory.json"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def _resolve_output_path(path: str, default_name: str) -> Path:
    """Keep generated files under data/outputs so the WebUI can serve them."""
    candidate = Path(path).expanduser() if path else OUTPUT_DIR / default_name
    if candidate.is_absolute():
        try:
            candidate.resolve().relative_to(DATA_DIR.resolve())
        except Exception:
            candidate = OUTPUT_DIR / candidate.name
    else:
        candidate = OUTPUT_DIR / candidate.name
    candidate.parent.mkdir(parents=True, exist_ok=True)
    return candidate

try:
    import certifi
except Exception:
    certifi = None
else:
    os.environ.setdefault("SSL_CERT_FILE", certifi.where())
    os.environ.setdefault("REQUESTS_CA_BUNDLE", certifi.where())

try:
    import truststore
except Exception:
    truststore = None
else:
    try:
        truststore.inject_into_ssl()
    except Exception:
        pass

# Google 官方 API 客戶端組件
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google.auth.exceptions import RefreshError
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

# DuckDuckGo 免金鑰搜尋
try:
    from ddgs import DDGS
except ImportError:
    from duckduckgo_search import DDGS

# =====================================================================
# GOOGLE OAUTH 憑證管理狀態機
# =====================================================================

GOOGLE_SCOPES = [
    'https://www.googleapis.com/auth/gmail.modify',
    'https://www.googleapis.com/auth/calendar',
    'https://www.googleapis.com/auth/drive',
    'https://www.googleapis.com/auth/spreadsheets'
]

def get_google_service(service_name: str, version: str):
    """動態安全獲取 Google 授權服務物件，內建過期自動刷新機制。"""
    creds = None

    def _google_request():
        if certifi is None:
            return Request()
        try:
            import requests
            session = requests.Session()
            session.verify = certifi.where()
            return Request(session=session)
        except Exception:
            return Request()

    BASE_DIR = Path(__file__).resolve().parent.parent.parent
    token_path = BASE_DIR / os.environ.get("GOOGLE_TOKEN_PATH", "token.json")
    creds_path = BASE_DIR / os.environ.get("GOOGLE_CREDENTIALS_PATH", "credentials.json")

    if token_path.exists():
        try:
            creds = Credentials.from_authorized_user_file(str(token_path), GOOGLE_SCOPES)
        except Exception:
            creds = None

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(_google_request())
            except RefreshError:
                if creds_path.exists():
                    flow = InstalledAppFlow.from_client_secrets_file(str(creds_path), GOOGLE_SCOPES)
                    creds = flow.run_local_server(port=0)
                else:
                    raise FileNotFoundError(f"找不到憑證檔案 {creds_path}，請確認檔案位置。")
        else:
            if creds_path.exists():
                flow = InstalledAppFlow.from_client_secrets_file(str(creds_path), GOOGLE_SCOPES)
                creds = flow.run_local_server(port=0)
            else:
                raise FileNotFoundError(f"找不到憑證檔案 {creds_path}，請確認檔案位置。")

        token_path.write_text(creds.to_json(), encoding="utf-8")

    return build(service_name, version, credentials=creds)

# =====================================================================
# 基礎核心工具
# =====================================================================

@tool()
def calculator(expression: str) -> dict:
    """執行數學運算，支援加減乘除與次方，防止任意 code 執行。"""
    allowed_names = {k: v for k, v in math.__dict__.items() if not k.startswith("_")}
    allowed_names.update({"abs": abs, "round": round, "int": int, "float": float})
    try:
        result = eval(expression, {"__builtins__": {}}, allowed_names)  # noqa: S307
        return {"result": result, "expression": expression}
    except ZeroDivisionError:
        return {"error": "除以零", "expression": expression}
    except Exception as e:
        return {"error": f"無效的表達式：{e}", "expression": expression}

@tool()
def get_current_time() -> dict:
    """取得目前的日期與時間（台灣時區 UTC+8）"""
    tz = datetime.timezone(datetime.timedelta(hours=8))
    now = datetime.datetime.now(tz)
    return {
        "datetime": now.strftime("%Y-%m-%d %H:%M:%S"),
        "weekday": ["一", "二", "三", "四", "五", "六", "日"][now.weekday()],
        "timezone": "Asia/Taipei (UTC+8)"
    }

@tool()
def read_file(path: str) -> dict:
    """讀取本地檔案的內容"""
    try:
        p = Path(path)
        if not p.exists(): return {"error": f"檔案不存在：{path}"}
        return {"content": p.read_text(encoding="utf-8")}
    except Exception as e:
        return {"error": str(e)}

@tool(schema_override={"properties": {"mode": {"enum": ["overwrite", "append"]}}})
def write_file(path: str, content: str, mode: str = "overwrite") -> dict:
    """將內容寫入本地檔案（若檔案不存在則建立）"""
    try:
        p = Path(path)
        p.parent.mkdir(parents=True, exist_ok=True)
        write_mode = "a" if mode == "append" else "w"
        with open(p, write_mode, encoding="utf-8") as f:
            f.write(content)
        return {"success": True, "path": str(p.resolve()), "mode": mode}
    except Exception as e:
        return {"error": str(e)}

@tool()
def http_get(url: str) -> dict:
    """發送 HTTP GET 請求並回傳回應內容（純文字，截斷至 2000 字元）"""
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Agent/1.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            text = resp.read().decode("utf-8", errors="ignore")
            return {"status": resp.status, "body": text[:2000], "truncated": len(text) > 2000}
    except Exception as e:
        return {"error": str(e), "url": url}

# =====================================================================
# Windows 系統呼叫
# =====================================================================

@tool()
def run_command(command: str) -> dict:
    """執行 Windows cmd 指令，回傳輸出結果"""
    try:
        res = subprocess.run(command, shell=True, capture_output=True, text=True, timeout=15)
        return {"stdout": res.stdout, "stderr": res.stderr, "returncode": res.returncode}
    except subprocess.TimeoutExpired:
        return {"error": "指令執行逾時"}

@tool()
def run_powershell(script: str) -> dict:
    """執行 PowerShell 腳本"""
    try:
        res = subprocess.run(["powershell", "-Command", script], capture_output=True, text=True, timeout=15)
        return {"stdout": res.stdout, "stderr": res.stderr, "returncode": res.returncode}
    except subprocess.TimeoutExpired:
        return {"error": "PowerShell 腳本執行逾時"}

@tool()
def list_processes(max_items: int = 80) -> dict:
    """列出目前 Windows 執行中的程序，並回傳容易閱讀的摘要。"""
    res = run_command("tasklist /FO CSV")
    if res.get("error"):
        return res
    rows = []
    try:
        reader = csv.DictReader((res.get("stdout") or "").splitlines())
        for row in reader:
            name = row.get("Image Name") or row.get("映像名稱") or ""
            pid = row.get("PID") or ""
            mem = row.get("Mem Usage") or row.get("記憶體使用量") or ""
            if name:
                rows.append({"name": name, "pid": pid, "memory": mem})
    except Exception:
        return res

    counts = {}
    for row in rows:
        counts[row["name"]] = counts.get(row["name"], 0) + 1
    top = sorted(counts.items(), key=lambda item: (-item[1], item[0].lower()))[:10]
    browser_names = {"chrome.exe", "msedge.exe", "firefox.exe", "brave.exe", "opera.exe"}
    browsers = [{"name": name, "count": count} for name, count in top if name.lower() in browser_names]
    limit = max(1, min(int(max_items or 80), 300))
    return {
        "success": res.get("returncode") == 0,
        "process_count": len(rows),
        "top_processes": [{"name": name, "count": count} for name, count in top],
        "browser_processes": browsers,
        "processes": rows[:limit],
        "truncated": len(rows) > limit,
        "returncode": res.get("returncode"),
        "stderr": res.get("stderr", ""),
    }

@tool()
def kill_process(process_name_or_pid: str) -> dict:
    """強制終止指定程序（PID 或名稱）"""
    cmd = f"taskkill /F /PID {process_name_or_pid}" if process_name_or_pid.isdigit() else f"taskkill /F /IM {process_name_or_pid}"
    return run_command(cmd)

@tool()
def get_env_variable(name: str) -> dict:
    """讀取系統或使用者環境變數"""
    val = os.environ.get(name)
    return {"name": name, "value": val} if val is not None else {"error": f"找不到變數 {name}"}

@tool()
def set_env_variable(name: str, value: str) -> dict:
    """設定環境變數（僅限於目前 Python 程序的生命週期）"""
    os.environ[name] = value
    return {"success": True, "name": name, "value": value}

@tool()
def list_directory(path: str) -> dict:
    """列出目錄內容（含大小、類型）"""
    try:
        items = []
        for entry in os.scandir(path):
            items.append({"name": entry.name, "is_file": entry.is_file(), "size": entry.stat().st_size if entry.is_file() else 0})
        return {"path": path, "items": items}
    except Exception as e:
        return {"error": str(e)}

@tool()
def move_file(src: str, dst: str) -> dict:
    """移動或重新命名檔案"""
    try:
        moved = shutil.move(src, dst)
        p = Path(moved)
        return {"success": True, "src": src, "dst": dst, "path": str(p.resolve()), "filename": p.name, "size": p.stat().st_size if p.is_file() else 0}
    except Exception as e:
        return {"error": str(e)}

@tool()
def delete_file(path: str) -> dict:
    """刪除檔案或空目錄"""
    try:
        p = Path(path)
        if p.is_file(): p.unlink()
        elif p.is_dir(): p.rmdir()
        return {"success": True}
    except Exception as e:
        return {"error": str(e)}

@tool()
def copy_file(src: str, dst: str) -> dict:
    """複製檔案至目標路徑"""
    try:
        copied = shutil.copy2(src, dst)
        p = Path(copied)
        return {"success": True, "src": src, "dst": dst, "path": str(p.resolve()), "filename": p.name, "size": p.stat().st_size if p.is_file() else 0}
    except Exception as e:
        return {"error": str(e)}

@tool()
def get_file_info(path: str) -> dict:
    """取得檔案屬性（大小、時間戳）"""
    try:
        stat = Path(path).stat()
        return {"size": stat.st_size, "modified": stat.st_mtime}
    except Exception as e:
        return {"error": str(e)}

@tool()
def clipboard_get() -> dict:
    """讀取目前 Windows 剪貼簿文字內容"""
    return run_powershell("Get-Clipboard")

@tool()
def clipboard_set(text: str) -> dict:
    """將文字寫入 Windows 剪貼簿"""
    escaped_text = text.replace("'", "''")
    return run_powershell(f"Set-Clipboard -Value '{escaped_text}'")

@tool()
def show_notification(title: str, message: str) -> dict:
    """顯示 Windows Toast 通知"""
    escaped_title = title.replace("'", "''")
    escaped_msg = message.replace("'", "''")
    ps = f"""
    [Windows.UI.Notifications.ToastNotificationManager, Windows.UI.Notifications, ContentType = WindowsRuntime] | Out-Null
    $template = [Windows.UI.Notifications.ToastNotificationManager]::GetTemplateContent([Windows.UI.Notifications.ToastTemplateType]::ToastText02)
    $textNodes = $template.GetElementsByTagName('text')
    $textNodes.Item(0).AppendChild($template.CreateTextNode('{escaped_title}')) | Out-Null
    $textNodes.Item(1).AppendChild($template.CreateTextNode('{escaped_msg}')) | Out-Null
    $toast = [Windows.UI.Notifications.ToastNotification]::new($template)
    [Windows.UI.Notifications.ToastNotificationManager]::CreateToastNotifier('AI Agent').Show($toast)
    """
    return run_powershell(ps)

@tool()
def open_url(url: str) -> dict:
    """用預設瀏覽器開啟 URL"""
    try:
        os.startfile(url)
        return {"success": True}
    except Exception as e:
        return {"error": str(e)}


def _now_iso() -> str:
    return datetime.datetime.now(datetime.timezone.utc).isoformat()


def _site_memory_load() -> dict:
    try:
        data = json.loads(SITE_MEMORY_FILE.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            data.setdefault("sites", [])
            data.setdefault("failures", [])
            return data
    except Exception:
        pass
    return {"sites": [], "failures": []}


def _site_memory_save(data: dict) -> None:
    SITE_MEMORY_FILE.parent.mkdir(parents=True, exist_ok=True)
    SITE_MEMORY_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _clean_url(url: str) -> str:
    raw = str(url or "").strip()
    if not raw:
        return ""
    parsed = urllib.parse.urlsplit(raw)
    if parsed.scheme and parsed.scheme.lower() not in {"http", "https"}:
        return ""
    if not parsed.scheme:
        raw = "https://" + raw
    parsed = urllib.parse.urlsplit(raw)
    if parsed.scheme.lower() not in {"http", "https"} or not parsed.netloc:
        return ""
    path = parsed.path or "/"
    if path != "/" and path.endswith("/"):
        path = path[:-1]
    return urllib.parse.urlunsplit((parsed.scheme.lower(), parsed.netloc.lower(), path, "", ""))


def _site_terms(text: str) -> set[str]:
    text = (text or "").lower()
    terms = {t for t in re.findall(r"[a-z0-9][a-z0-9._-]{1,}", text)}
    institution_terms: set[str] = set()
    for m in re.finditer(
        r"(?:我是|我就讀|使用者是|就讀|來自|在|到)?([\u4e00-\u9fff]{1,8}?(?:高級中學|高中|中學|大學|國中|國小|學院|學校|公司))",
        text,
    ):
        institution_terms.add(m.group(1))
    for m in re.finditer(r"(?:我是|使用者是)?([\u4e00-\u9fff]{2,8}?)(?:學生|生|校友)", text):
        institution_terms.add(m.group(1))
    for chunk in re.findall(r"[\u4e00-\u9fff]{2,}", text):
        if re.search(r"(使用者|任務|查詢|進度|請|幫|帮|登入|作業|課程|網站|平台|學生|校友)", chunk):
            continue
        if len(chunk) <= 8:
            terms.add(chunk)
    for term in institution_terms:
        terms.add(term)
        for suffix, short in (("中學", "中"), ("高中", "高"), ("大學", "大"), ("國中", "中"), ("國小", "小")):
            if term.endswith(suffix) and len(term) > len(suffix):
                terms.add(term[0] + short)
    service_aliases = {
        "moodle": {"數位學習", "教學平台", "learning"},
        "classroom": {"課程", "教室", "learning"},
        "portal": {"入口", "平台"},
        "lms": {"數位學習", "教學平台", "learning"},
    }
    for service, aliases in service_aliases.items():
        if service in text:
            terms.update(aliases)
    return {t for t in terms if t}


_GENERIC_SITE_TERMS = {
    "moodle", "learning", "login", "portal", "course", "courses",
    "網站", "平台", "登入", "入口", "網址", "教學平台", "數位學習",
}


def _required_site_terms(query: str) -> set[str]:
    return {t for t in _site_terms(query) if t not in _GENERIC_SITE_TERMS}


def _site_score(query: str, record: dict) -> float:
    query_terms = _site_terms(query)
    hay = " ".join(str(record.get(k, "")) for k in ("service", "url", "title", "note"))
    hay_terms = _site_terms(hay)
    if not query_terms:
        return 0.0
    required = _required_site_terms(query)
    if required and not (required & hay_terms):
        return -1.0
    overlap = len(query_terms & hay_terms) / max(1, len(query_terms))
    if required:
        overlap += 0.5 * (len(required & hay_terms) / max(1, len(required)))
    q_norm = re.sub(r"\s+", "", query.lower())
    h_norm = re.sub(r"\s+", "", hay.lower())
    if q_norm and q_norm in h_norm:
        overlap = max(overlap, 0.9)
    return min(1.0, overlap)


def _site_entry_quality(record: dict, query: str = "") -> float:
    hay = " ".join(str(record.get(k, "")) for k in ("service", "url", "title", "note")).lower()
    url = _clean_url(str(record.get("url", "")))
    parsed = urllib.parse.urlsplit(url)
    path = parsed.path.lower() or "/"
    score = 0.0
    if re.search(r"(^|\s)(錯誤|error|invalid|失敗)(\s|$)", hay, re.I):
        score -= 0.6
    if re.search(r"(/auth/|oauth|callback|forgot|calendar|詳細月曆|archive|archiv|pluginfile|/mod/)", path + " " + hay, re.I):
        score -= 0.25
    if re.search(r"(\d{2,4}\s*學年|\b(?:19|20)\d{2}\b)", hay, re.I):
        score -= 0.2
    if path == "/" or re.search(r"^/(login|signin|home|index)(/|\.|$)", path, re.I):
        score += 0.25
    required = _required_site_terms(query)
    if required and required & _site_terms(hay):
        score += 0.15
    return score


def _site_record_is_usable_success(record: dict, query: str = "") -> bool:
    return _site_entry_quality(record, query) > -0.3


def _site_memory_remember(service: str, url: str, title: str = "", note: str = "", status: str = "success") -> dict:
    clean_url = _clean_url(url)
    if not clean_url:
        return {"error": "缺少 URL"}
    candidate_record = {"service": service, "url": clean_url, "title": title, "note": note}
    if (status or "success") == "success" and not _site_record_is_usable_success(candidate_record, service):
        return _site_memory_failure(
            service,
            clean_url,
            "頁面看起來像錯誤頁、驗證中繼頁或非入口頁",
            note or title,
        )
    data = _site_memory_load()
    service = str(service or title or urllib.parse.urlsplit(clean_url).netloc).strip()
    title = str(title or "").strip()
    note = str(note or "").strip()
    found = None
    for site in data["sites"]:
        if _clean_url(site.get("url", "")) == clean_url:
            found = site
            break
    if not found:
        found = {
            "service": service,
            "url": clean_url,
            "title": title,
            "note": note,
            "status": status or "success",
            "success_count": 0,
            "created_at": _now_iso(),
        }
        data["sites"].append(found)
    found["service"] = service or found.get("service", "")
    if title:
        found["title"] = title
    if note:
        old_note = found.get("note", "")
        found["note"] = note if note in old_note else (old_note + " | " + note).strip(" |")
    found["status"] = status or found.get("status", "success")
    found["success_count"] = int(found.get("success_count", 0) or 0) + (1 if found["status"] == "success" else 0)
    found["last_seen_at"] = _now_iso()
    _site_memory_save(data)
    return {"success": True, "site": found}


def _site_memory_failure(service: str, url: str, error: str = "", note: str = "") -> dict:
    clean_url = _clean_url(url)
    if not clean_url:
        return {"error": "缺少 URL"}
    data = _site_memory_load()
    service = str(service or urllib.parse.urlsplit(clean_url).netloc).strip()
    found = None
    for item in data["failures"]:
        if _clean_url(item.get("url", "")) == clean_url:
            found = item
            break
    if not found:
        found = {"service": service, "url": clean_url, "created_at": _now_iso(), "count": 0}
        data["failures"].append(found)
    found.update({
        "service": service or found.get("service", ""),
        "error": str(error or "")[:500],
        "note": str(note or "")[:500],
        "last_failed_at": _now_iso(),
        "count": int(found.get("count", 0) or 0) + 1,
    })
    _site_memory_save(data)
    return {"success": True, "failure": found}


@tool()
def site_memory_search(query: str, max_results: int = 5) -> dict:
    """搜尋 Raphael 已學過的網站入口與失敗網址，不包含帳密。"""
    data = _site_memory_load()
    limit = max(1, min(int(max_results or 5), 20))
    sites = []
    for site in data.get("sites", []):
        score = _site_score(query, site)
        if score > 0 or not query:
            row = dict(site)
            row["score"] = round(score, 4)
            sites.append(row)
    sites.sort(
        key=lambda row: (
            row.get("score", 0),
            _site_entry_quality(row, query),
            int(row.get("success_count", 0) or 0),
        ),
        reverse=True,
    )
    failures = []
    for failure in data.get("failures", []):
        score = _site_score(query, failure)
        if score > 0 or not query:
            row = dict(failure)
            row["score"] = round(score, 4)
            failures.append(row)
    failures.sort(key=lambda row: (row.get("score", 0), int(row.get("count", 0) or 0)), reverse=True)
    return {
        "success": True,
        "sites": sites[:limit],
        "failures": failures[:limit],
        "site_count": len(sites),
        "failure_count": len(failures),
    }


@tool()
def site_memory_remember(service: str, url: str, title: str = "", note: str = "") -> dict:
    """記住已驗證可用的網站入口；只存服務、網址、標題與備註，不存帳密。"""
    return _site_memory_remember(service, url, title, note, "success")


@tool()
def site_memory_mark_failure(service: str, url: str, error: str = "", note: str = "") -> dict:
    """記住連不上、DNS 失敗或確認錯誤的網站入口，避免下次重試同樣錯誤。"""
    return _site_memory_failure(service, url, error, note)


def _probe_website(url: str, timeout: int = 8) -> dict:
    clean = _clean_url(url)
    if not clean:
        return {"ok": False, "url": url, "error": "empty url"}
    try:
        req = urllib.request.Request(
            clean,
            headers={
                "User-Agent": "Mozilla/5.0 RaphaelSiteResolver/1.0",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            },
        )
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read(400_000)
            text = raw.decode(resp.headers.get_content_charset() or "utf-8", errors="replace")
            title_match = re.search(r"<title[^>]*>(.*?)</title>", text, flags=re.I | re.S)
            title = re.sub(r"\s+", " ", title_match.group(1)).strip() if title_match else ""
            return {
                "ok": True,
                "url": clean,
                "final_url": resp.geturl(),
                "status": getattr(resp, "status", 200),
                "title": title,
                "content_type": resp.headers.get("content-type", ""),
            }
    except Exception as e:
        return {"ok": False, "url": clean, "error": str(e)}


def _candidate_score(query: str, candidate: dict, preferred_domain: str = "", must_contain: str = "") -> float:
    hay = " ".join(str(candidate.get(k, "")) for k in ("title", "href", "url", "body", "final_url"))
    q_terms = _site_terms(query)
    hay_terms = _site_terms(hay)
    required = _required_site_terms(query)
    if required and not (required & hay_terms):
        return -1.0
    score = len(q_terms & hay_terms) / max(1, len(q_terms))
    if required:
        score += 0.5 * (len(required & hay_terms) / max(1, len(required)))
    if preferred_domain and preferred_domain.lower() in hay.lower():
        score += 0.5
    if must_contain and must_contain.lower() in hay.lower():
        score += 0.3
    if re.search(r"(忘記密碼|forgot|calendar|詳細月曆|archive|archiv|pluginfile|/mod/)", hay, re.I):
        score -= 0.3
    if re.search(r"(\d{2,4}\s*學年|\b(?:19|20)\d{2}\b)", hay, re.I):
        score -= 0.4
    score += _site_entry_quality(candidate, query)
    return min(1.0, score)


def _rank_website_candidates(query: str, candidates: list[dict], known_failures: set[str], preferred_domain: str = "", must_contain: str = "") -> list[dict]:
    deduped = {}
    required_terms = _required_site_terms(query)
    min_score = 0.2 if required_terms else 0.05
    for candidate in candidates:
        clean = _clean_url(candidate.get("url", ""))
        if not clean or clean in known_failures:
            continue
        candidate = dict(candidate)
        candidate["url"] = clean
        candidate["score"] = max(
            float(candidate.get("score", 0) or 0),
            _candidate_score(query, candidate, preferred_domain, must_contain),
        )
        if candidate["score"] < min_score:
            continue
        if clean not in deduped or candidate["score"] > deduped[clean]["score"]:
            deduped[clean] = candidate
    return sorted(deduped.values(), key=lambda c: c.get("score", 0), reverse=True)


def _verify_website_candidates(query: str, ranked: list[dict], limit: int, preferred_domain: str = "", must_contain: str = "") -> tuple[list[dict], list[dict]]:
    verified = []
    failures = []
    for candidate in ranked[:max(limit * 2, 8)]:
        probe = _probe_website(candidate["url"])
        candidate = {**candidate, **probe}
        if probe.get("ok"):
            candidate["score"] = max(
                float(candidate.get("score", 0) or 0),
                _candidate_score(query, {**candidate, "title": probe.get("title") or candidate.get("title", "")}, preferred_domain, must_contain),
            )
            verified.append(candidate)
        else:
            failures.append(candidate)
    verified.sort(key=lambda c: (c.get("score", 0), _site_entry_quality(c, query)), reverse=True)
    return verified, failures


@tool()
def website_find(query: str, preferred_domain: str = "", must_contain: str = "", max_results: int = 5) -> dict:
    """根據目標描述尋找並驗證網站入口，會先查已學入口，再搜尋網路並避開已知失敗網址。"""
    limit = max(1, min(int(max_results or 5), 10))
    memory = site_memory_search(query, max_results=limit)
    known_failures = {_clean_url(f.get("url", "")) for f in memory.get("failures", [])}
    memory_candidates = []
    for site in memory.get("sites", []):
        if _clean_url(site.get("url", "")) not in known_failures:
            memory_candidates.append({
                "source": "site_memory",
                "url": site.get("url", ""),
                "title": site.get("title", ""),
                "score": site.get("score", 0),
            })
    memory_ranked = _rank_website_candidates(query, memory_candidates, known_failures, preferred_domain, must_contain)
    high_confidence_memory = [candidate for candidate in memory_ranked if float(candidate.get("score", 0) or 0) >= 0.65]
    if high_confidence_memory:
        verified, failures = _verify_website_candidates(query, high_confidence_memory, limit, preferred_domain, must_contain)
        best = verified[0] if verified else None
        if best:
            _site_memory_remember(
                query,
                best.get("final_url") or best.get("url"),
                best.get("title", ""),
                "website_find memory verified",
                "success",
            )
            return {
                "success": True,
                "best": best,
                "verified": verified[:limit],
                "failed_candidates": failures[:limit],
                "searched": False,
                "search_error": "",
                "memory_first": True,
            }

    candidates = list(memory_candidates)
    search = web_search(query, max_results=max(limit * 2, 8))
    if isinstance(search, dict) and isinstance(search.get("results"), list):
        for item in search["results"]:
            if not isinstance(item, dict):
                continue
            href = item.get("href") or item.get("url") or item.get("link")
            if not href:
                continue
            candidates.append({
                "source": "web_search",
                "url": href,
                "title": item.get("title", ""),
                "body": item.get("body", "") or item.get("snippet", ""),
            })

    ranked = _rank_website_candidates(query, candidates, known_failures, preferred_domain, must_contain)
    verified, failures = _verify_website_candidates(query, ranked, limit, preferred_domain, must_contain)
    best = verified[0] if verified else None
    if best:
        _site_memory_remember(
            query,
            best.get("final_url") or best.get("url"),
            best.get("title", ""),
            "website_find verified",
            "success",
        )
    return {
        "success": bool(best),
        "best": best,
        "verified": verified[:limit],
        "failed_candidates": failures[:limit],
        "searched": "error" not in search,
        "search_error": search.get("error") if isinstance(search, dict) else "",
    }


_BROWSER_STATE = {
    "playwright": None,
    "browser": None,
    "context": None,
    "page": None,
    "headless": None,
}


def _close_background_browser() -> None:
    browser = _BROWSER_STATE.get("browser")
    context = _BROWSER_STATE.get("context")
    playwright = _BROWSER_STATE.get("playwright")
    with contextlib.suppress(Exception):
        if context:
            context.close()
    with contextlib.suppress(Exception):
        if browser:
            browser.close()
    with contextlib.suppress(Exception):
        if playwright:
            playwright.stop()
    _BROWSER_STATE.update({"playwright": None, "browser": None, "context": None, "page": None, "headless": None})


atexit.register(_close_background_browser)


def _background_browser(headless: bool = True, width: int = 1280, height: int = 900):
    try:
        from playwright.sync_api import sync_playwright
    except Exception as e:
        return None, (
            "背景瀏覽器工具需要 playwright。請執行："
            "python -m pip install playwright && python -m playwright install chromium。"
            f" 原始錯誤：{e}"
        )

    if _BROWSER_STATE["context"] is not None and _BROWSER_STATE["headless"] == bool(headless):
        page = _BROWSER_STATE["page"]
        if page is None or page.is_closed():
            page = _BROWSER_STATE["context"].new_page()
            _BROWSER_STATE["page"] = page
        return page, None

    if _BROWSER_STATE["context"] is not None:
        _close_background_browser()

    profile_dir = DATA_DIR / "browser_profile"
    profile_dir.mkdir(parents=True, exist_ok=True)
    errors = []
    try:
        pw = sync_playwright().start()
        for launch_args in (
            {"channel": "msedge"},
            {"channel": "chrome"},
            {},
        ):
            try:
                context = pw.chromium.launch_persistent_context(
                    str(profile_dir),
                    headless=bool(headless),
                    viewport={"width": int(width), "height": int(height)},
                    args=["--disable-blink-features=AutomationControlled"],
                    **launch_args,
                )
                page = context.pages[0] if context.pages else context.new_page()
                page.set_default_timeout(10_000)
                _BROWSER_STATE.update({
                    "playwright": pw,
                    "browser": None,
                    "context": context,
                    "page": page,
                    "headless": bool(headless),
                })
                return page, None
            except Exception as e:
                errors.append(f"{launch_args or 'bundled chromium'}: {e}")
                try:
                    browser = pw.chromium.launch(
                        headless=bool(headless),
                        args=["--disable-blink-features=AutomationControlled"],
                        **launch_args,
                    )
                    context = browser.new_context(viewport={"width": int(width), "height": int(height)})
                    page = context.new_page()
                    page.set_default_timeout(10_000)
                    _BROWSER_STATE.update({
                        "playwright": pw,
                        "browser": browser,
                        "context": context,
                        "page": page,
                        "headless": bool(headless),
                    })
                    return page, None
                except Exception as fallback_e:
                    errors.append(f"{launch_args or 'bundled chromium'} transient: {fallback_e}")
        with contextlib.suppress(Exception):
            pw.stop()
    except Exception as e:
        errors.append(str(e))
    return None, "背景瀏覽器啟動失敗：" + " | ".join(errors[-3:])


def _browser_screenshot_result(page, name_prefix: str = "browser") -> dict:
    out = _resolve_output_path(
        "",
        f"{name_prefix}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.png",
    )
    page.screenshot(path=str(out), full_page=True)
    return _file_result(out, mime="image/png")


def _browser_controls(page, limit: int = 60) -> list[dict]:
    script = """
    (limit) => {
      const labelFor = (el) => {
        if (el.id) {
          const label = document.querySelector(`label[for="${CSS.escape(el.id)}"]`);
          if (label && label.innerText) return label.innerText.trim();
        }
        const parent = el.closest('label');
        if (parent && parent.innerText) return parent.innerText.trim();
        return '';
      };
      const cssFor = (el) => {
        if (el.id) return `#${CSS.escape(el.id)}`;
        if (el.name) return `${el.tagName.toLowerCase()}[name="${CSS.escape(el.name)}"]`;
        const type = el.getAttribute('type');
        if (type) return `${el.tagName.toLowerCase()}[type="${CSS.escape(type)}"]`;
        return el.tagName.toLowerCase();
      };
      return Array.from(document.querySelectorAll('input, textarea, select, button, a, [role="button"]'))
        .filter(el => {
          const r = el.getBoundingClientRect();
          const style = window.getComputedStyle(el);
          return r.width > 0 && r.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
        })
        .slice(0, limit)
        .map(el => ({
          selector: cssFor(el),
          tag: el.tagName.toLowerCase(),
          type: el.getAttribute('type') || '',
          name: el.getAttribute('name') || '',
          placeholder: el.getAttribute('placeholder') || '',
          label: labelFor(el),
          text: (el.innerText || el.value || el.getAttribute('aria-label') || '').trim().slice(0, 120),
          href: el.href || '',
        }));
    }
    """
    try:
        return page.evaluate(script, int(limit))
    except Exception:
        return []


def _browser_links(page, query: str = "", limit: int = 80) -> list[dict]:
    script = """
    ({ query, limit }) => {
      const visible = (el) => {
        const r = el.getBoundingClientRect();
        const style = window.getComputedStyle(el);
        return r.width > 0 && r.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
      };
      const q = String(query || '').trim().toLowerCase();
      const links = Array.from(document.querySelectorAll('a[href]'))
        .filter(visible)
        .map(el => ({
          el,
          text: (el.innerText || el.getAttribute('aria-label') || el.title || '').trim().replace(/\\s+/g, ' '),
          href: el.href || '',
        }))
        .filter(item => {
          if (!q) return true;
          return item.text.toLowerCase().includes(q) || item.href.toLowerCase().includes(q);
        })
        .slice(0, limit);
      links.forEach((item, index) => item.el.setAttribute('data-raphael-link-index', String(index)));
      return links.map((item, index) => ({
        index,
        text: item.text.slice(0, 160),
        href: item.href,
        selector: `[data-raphael-link-index="${index}"]`,
      }));
    }
    """
    try:
        return page.evaluate(script, {"query": query or "", "limit": int(limit or 80)})
    except Exception:
        return []


def _browser_summary(page, *, screenshot: bool = False, text_limit: int = 4500) -> dict:
    try:
        body_text = page.locator("body").inner_text(timeout=2500)
    except Exception:
        body_text = ""
    result = {
        "success": True,
        "url": page.url,
        "title": page.title(),
        "text": body_text[:text_limit],
        "text_truncated": len(body_text) > text_limit,
        "controls": _browser_controls(page),
    }
    lowered = body_text.lower()
    if any(k in lowered for k in ("驗證碼", "captcha", "二階段", "2-step", "verification code", "authenticator", "核准登入")):
        result["needs_user_action"] = True
        result["message"] = "頁面似乎需要驗證碼、二階段驗證或使用者本人確認。"
    if screenshot:
        result["screenshot"] = _browser_screenshot_result(page)
    return result


def _first_visible_locator(page, selectors: list[str]):
    for selector in selectors:
        try:
            loc = page.locator(selector).first
            if loc.count() and loc.is_visible():
                return loc
        except Exception:
            continue
    return None


def _fill_first(page, selectors: list[str], value: str) -> bool:
    loc = _first_visible_locator(page, selectors)
    if not loc:
        return False
    loc.fill(value)
    return True


def _click_text(page, labels: list[str]) -> bool:
    for label in labels:
        for getter in (
            lambda: page.get_by_role("button", name=re.compile(label, re.I)).first,
            lambda: page.get_by_role("link", name=re.compile(label, re.I)).first,
            lambda: page.get_by_text(re.compile(label, re.I)).first,
        ):
            try:
                loc = getter()
                if loc.count() and loc.is_visible():
                    loc.click()
                    page.wait_for_load_state("domcontentloaded", timeout=8000)
                    return True
            except Exception:
                continue
    return False


def _remember_browser_page(page, note: str = "background browser visit") -> None:
    with contextlib.suppress(Exception):
        title = page.title()
        url = page.url
        if _clean_url(url):
            service = title or urllib.parse.urlsplit(url).netloc
            _site_memory_remember(service, url, title, note, "success")


@tool()
def browser_open(url: str, headless: bool = True, screenshot: bool = False) -> dict:
    """在獨立背景瀏覽器開啟 URL，不干擾使用者目前操作的瀏覽器。"""
    page, error = _background_browser(headless=headless)
    if error:
        return {"error": error}
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30_000)
        with contextlib.suppress(Exception):
            page.wait_for_load_state("networkidle", timeout=5_000)
        _remember_browser_page(page, "browser_open succeeded")
        return _browser_summary(page, screenshot=screenshot)
    except Exception as e:
        _site_memory_failure("", url, str(e), "browser_open failed")
        return {"error": str(e), "url": url}


@tool()
def browser_get_page(screenshot: bool = False) -> dict:
    """讀取背景瀏覽器目前頁面的文字、可操作元件與選擇器。"""
    page, error = _background_browser(headless=True)
    if error:
        return {"error": error}
    return _browser_summary(page, screenshot=screenshot)


@tool()
def browser_links(query: str = "", limit: int = 80) -> dict:
    """列出背景瀏覽器目前頁面的可見連結；可用 query 過濾文字或網址，回傳 index 供 browser_follow_link 使用。"""
    page, error = _background_browser(headless=True)
    if error:
        return {"error": error}
    links = _browser_links(page, query=query, limit=limit)
    return {
        "success": True,
        "url": page.url,
        "title": page.title(),
        "query": query,
        "links": links,
        "count": len(links),
    }


@tool()
def browser_follow_link(index: int, screenshot: bool = False) -> dict:
    """依照 browser_links 回傳的 index 點擊背景瀏覽器中的連結。"""
    page, error = _background_browser(headless=True)
    if error:
        return {"error": error}
    try:
        locator = page.locator(f'[data-raphael-link-index="{int(index)}"]').first
        if not locator.count() or not locator.is_visible():
            return {"error": f"找不到連結編號：{index}", "links": _browser_links(page, limit=30), **_browser_summary(page, screenshot=True)}
        locator.click()
        with contextlib.suppress(Exception):
            page.wait_for_load_state("domcontentloaded", timeout=8_000)
        with contextlib.suppress(Exception):
            page.wait_for_load_state("networkidle", timeout=4_000)
        return _browser_summary(page, screenshot=screenshot)
    except Exception as e:
        return {"error": str(e), "index": index, **_browser_summary(page, screenshot=True)}


@tool()
def browser_click(target: str, screenshot: bool = False) -> dict:
    """在背景瀏覽器點擊 CSS selector、文字、連結或按鈕。"""
    page, error = _background_browser(headless=True)
    if error:
        return {"error": error}
    try:
        if target.startswith("css="):
            page.locator(target[4:]).first.click()
        else:
            clicked = False
            with contextlib.suppress(Exception):
                loc = page.locator(target).first
                if loc.count() and loc.is_visible():
                    loc.click()
                    clicked = True
            if not clicked:
                clicked = _click_text(page, [target])
            if not clicked:
                return {"error": f"找不到可點擊目標：{target}", **_browser_summary(page, screenshot=True)}
        with contextlib.suppress(Exception):
            page.wait_for_load_state("domcontentloaded", timeout=8_000)
        return _browser_summary(page, screenshot=screenshot)
    except Exception as e:
        return {"error": str(e), "target": target, **_browser_summary(page, screenshot=True)}


@tool()
def browser_fill(target: str, text: str = "", secret_text: str = "", press_enter: bool = False) -> dict:
    """在背景瀏覽器填入欄位；密碼等敏感值請放 secret_text。target 可為 CSS selector、label、placeholder 或欄位名稱。"""
    page, error = _background_browser(headless=True)
    if error:
        return {"error": error}
    value = secret_text if secret_text else text
    if not value:
        return {"error": "缺少要填入的文字"}
    selectors = []
    if target.startswith("css="):
        selectors.append(target[4:])
    else:
        safe = target.replace('"', '\\"')
        selectors.extend([
            target,
            f'input[name="{safe}"]',
            f'textarea[name="{safe}"]',
            f'input[placeholder*="{safe}"]',
            f'textarea[placeholder*="{safe}"]',
            f'input[id="{safe}"]',
        ])
    try:
        filled = _fill_first(page, selectors, value)
        if not filled:
            for getter in (
                lambda: page.get_by_label(re.compile(target, re.I)).first,
                lambda: page.get_by_placeholder(re.compile(target, re.I)).first,
            ):
                with contextlib.suppress(Exception):
                    loc = getter()
                    if loc.count() and loc.is_visible():
                        loc.fill(value)
                        filled = True
                        break
        if not filled:
            return {"error": f"找不到可填入欄位：{target}", **_browser_summary(page, screenshot=True)}
        if press_enter:
            page.keyboard.press("Enter")
            with contextlib.suppress(Exception):
                page.wait_for_load_state("domcontentloaded", timeout=8_000)
        return _browser_summary(page, screenshot=False)
    except Exception as e:
        return {"error": str(e), "target": target, **_browser_summary(page, screenshot=True)}


@tool(schema_override={"properties": {"key": {"enum": ["Enter", "Tab", "Escape", "ArrowDown", "ArrowUp", "Space"]}}})
def browser_press_key(key: str = "Enter") -> dict:
    """在背景瀏覽器送出常用按鍵。"""
    page, error = _background_browser(headless=True)
    if error:
        return {"error": error}
    try:
        page.keyboard.press(key)
        with contextlib.suppress(Exception):
            page.wait_for_load_state("domcontentloaded", timeout=8_000)
        return _browser_summary(page)
    except Exception as e:
        return {"error": str(e), "key": key}


@tool(schema_override={"properties": {"state": {"enum": ["none", "domcontentloaded", "load", "networkidle"]}}})
def browser_wait(milliseconds: int = 1000, state: str = "none", screenshot: bool = False) -> dict:
    """讓背景瀏覽器等待一段時間或等待頁面狀態穩定。"""
    page, error = _background_browser(headless=True)
    if error:
        return {"error": error}
    try:
        delay = max(0, min(int(milliseconds or 0), 30_000))
        if delay:
            page.wait_for_timeout(delay)
        if state in {"domcontentloaded", "load", "networkidle"}:
            with contextlib.suppress(Exception):
                page.wait_for_load_state(state, timeout=max(2_000, min(30_000, delay + 5_000)))
        return _browser_summary(page, screenshot=screenshot)
    except Exception as e:
        return {"error": str(e), "state": state, "milliseconds": milliseconds, **_browser_summary(page, screenshot=True)}


@tool()
def browser_back(screenshot: bool = False) -> dict:
    """讓背景瀏覽器回到上一頁。"""
    page, error = _background_browser(headless=True)
    if error:
        return {"error": error}
    try:
        response = page.go_back(wait_until="domcontentloaded", timeout=10_000)
        status = response.status if response else None
        summary = _browser_summary(page, screenshot=screenshot)
        summary["status"] = status
        return summary
    except Exception as e:
        return {"error": str(e), **_browser_summary(page, screenshot=True)}


@tool()
def browser_scroll(amount: int = 800, screenshot: bool = False) -> dict:
    """在背景瀏覽器滾動頁面；正數向下，負數向上。"""
    page, error = _background_browser(headless=True)
    if error:
        return {"error": error}
    try:
        page.mouse.wheel(0, int(amount))
        with contextlib.suppress(Exception):
            page.wait_for_timeout(400)
        return _browser_summary(page, screenshot=screenshot)
    except Exception as e:
        return {"error": str(e), "amount": amount, **_browser_summary(page, screenshot=True)}


@tool()
def browser_screenshot() -> dict:
    """擷取背景瀏覽器目前頁面截圖，回傳 WebUI 可開啟的檔案。"""
    page, error = _background_browser(headless=True)
    if error:
        return {"error": error}
    try:
        return {"success": True, **_browser_screenshot_result(page)}
    except Exception as e:
        return {"error": str(e)}


@tool()
def browser_login(url: str, username: str, password: str, use_google: bool = False, headless: bool = True) -> dict:
    """在獨立背景瀏覽器登入網站；支援一般帳密欄位與常見 Google SSO 流程。"""
    page, error = _background_browser(headless=headless)
    if error:
        return {"error": error}
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30_000)
        with contextlib.suppress(Exception):
            page.wait_for_load_state("networkidle", timeout=5_000)

        if use_google or not _first_visible_locator(page, ["input[type='password']", "input[name='password']", "#password"]):
            _click_text(page, ["google", "使用.*google", "google.*登入", "log in", "login", "登入"])

        user_ok = _fill_first(page, [
            "input[type='email']",
            "input[name='identifier']",
            "#identifierId",
            "input[name='username']",
            "input[name='email']",
            "input[type='text']",
        ], username)
        if user_ok:
            if _click_text(page, ["下一步", "next", "繼續", "continue"]):
                with contextlib.suppress(Exception):
                    page.wait_for_timeout(800)
            elif not _first_visible_locator(page, ["input[type='password']", "input[name='password']", "#password"]):
                page.keyboard.press("Enter")
                with contextlib.suppress(Exception):
                    page.wait_for_load_state("domcontentloaded", timeout=8_000)

        pass_ok = _fill_first(page, [
            "input[type='password']",
            "input[name='password']",
            "#password",
            "input[name='Passwd']",
        ], password)
        if not pass_ok:
            summary = _browser_summary(page, screenshot=True)
            summary.update({"logged_in": False, "message": "找不到密碼欄位，可能已跳到驗證頁或登入方式不同。"})
            return summary

        if not _click_text(page, ["登入", "login", "sign in", "下一步", "next", "繼續", "continue"]):
            page.keyboard.press("Enter")
        with contextlib.suppress(Exception):
            page.wait_for_load_state("networkidle", timeout=12_000)
        with contextlib.suppress(Exception):
            page.wait_for_timeout(1200)

        summary = _browser_summary(page, screenshot=True)
        text_lower = (summary.get("text") or "").lower()
        failed = any(k in text_lower for k in ("incorrect", "invalid password", "密碼錯誤", "登入失敗", "無效"))
        still_password = _first_visible_locator(page, [
            "input[type='password']",
            "input[name='password']",
            "#password",
            "input[name='Passwd']",
        ]) is not None
        still_login_url = bool(re.search(r"/login|signin|sign-in|accounts\.google", page.url, re.I))
        success_markers = any(k in text_lower for k in ("登出", "logout", "dashboard", "my courses", "我的課程", "個人檔案"))
        summary["logged_in"] = (
            not failed
            and not summary.get("needs_user_action", False)
            and not still_password
            and (success_markers or not still_login_url)
        )
        if failed:
            summary["message"] = "登入送出後頁面顯示帳號或密碼可能不正確。"
        elif summary.get("needs_user_action"):
            summary["message"] = "已送出帳密，但頁面需要使用者完成額外驗證。"
        elif not summary["logged_in"]:
            summary["message"] = "登入送出後仍停留在登入頁或密碼欄位仍可見，尚未確認登入成功。"
        else:
            summary["message"] = "已在背景瀏覽器送出登入流程。"
            _remember_browser_page(page, "browser_login completed")
        return summary
    except Exception as e:
        _site_memory_failure("", url, str(e), "browser_login failed")
        return {"error": str(e), "url": url, **_browser_summary(page, screenshot=True)}


@tool()
def browser_close() -> dict:
    """關閉背景瀏覽器工作階段。"""
    _close_background_browser()
    return {"success": True}

@tool(schema_override={"properties": {"hive": {"enum": ["HKEY_LOCAL_MACHINE", "HKEY_CURRENT_USER"]}}})
def registry_read(hive: str, key: str, value_name: str) -> dict:
    """讀取登錄檔指定機碼的值"""
    try:
        h = winreg.HKEY_LOCAL_MACHINE if hive == "HKEY_LOCAL_MACHINE" else winreg.HKEY_CURRENT_USER
        with winreg.OpenKey(h, key) as k:
            val, _ = winreg.QueryValueEx(k, value_name)
            return {"value": val}
    except Exception as e:
        return {"error": str(e)}

@tool(schema_override={"properties": {"hive": {"enum": ["HKEY_LOCAL_MACHINE", "HKEY_CURRENT_USER"]}}})
def registry_write(hive: str, key: str, value_name: str, value: str) -> dict:
    """寫入登錄檔機碼值（純字串格式）"""
    try:
        h = winreg.HKEY_LOCAL_MACHINE if hive == "HKEY_LOCAL_MACHINE" else winreg.HKEY_CURRENT_USER
        with winreg.OpenKey(h, key, 0, winreg.KEY_SET_VALUE) as k:
            winreg.SetValueEx(k, value_name, 0, winreg.REG_SZ, value)
            return {"success": True}
    except Exception as e:
        return {"error": str(e)}

# =====================================================================
# 額外實體 API (通訊, 開發者, 本地模型)
# =====================================================================

@tool()
def web_search(query: str, max_results: int = 3) -> dict:
    """使用 DuckDuckGo 進行網頁檢索，回傳標題與摘要"""
    errors = []
    def _isolated_ddgs_search() -> dict:
        code = (
            "import json, sys\n"
            "from ddgs import DDGS\n"
            "query = sys.argv[1]\n"
            "limit = int(sys.argv[2])\n"
            "client = DDGS(verify=False)\n"
            "errors = []\n"
            "for backend in ('brave','mojeek','wikipedia',None):\n"
            "    try:\n"
            "        kwargs = {'max_results': limit}\n"
            "        if backend: kwargs['backend'] = backend\n"
            "        results = list(client.text(query, **kwargs))\n"
            "        if results:\n"
            "            print(json.dumps({'results': results, 'backend': backend or 'auto'}, ensure_ascii=False))\n"
            "            raise SystemExit(0)\n"
            "    except Exception as e:\n"
            "        errors.append(f'{backend or \"auto\"}: {e}')\n"
            "print(json.dumps({'error': '搜尋沒有取得結果', 'attempts': errors[-4:]}, ensure_ascii=False))\n"
        )
        env = os.environ.copy()
        env.pop("SSL_CERT_FILE", None)
        env.pop("REQUESTS_CA_BUNDLE", None)
        res = subprocess.run(
            [sys.executable, "-c", code, query, str(max_results)],
            capture_output=True,
            text=True,
            timeout=30,
            env=env,
        )
        if res.returncode != 0 and not res.stdout.strip():
            return {"error": res.stderr.strip() or f"搜尋子程序失敗: {res.returncode}"}
        return json.loads(res.stdout)

    try:
        try:
            client = DDGS(verify=False)
        except TypeError:
            client = DDGS()
        for backend in ("brave", "mojeek", "wikipedia", None):
            try:
                kwargs = {"max_results": max_results}
                if backend:
                    kwargs["backend"] = backend
                results = list(client.text(query, **kwargs))
                if results:
                    return {"results": results, "backend": backend or "auto"}
            except Exception as e:
                errors.append(f"{backend or 'auto'}: {e}")
        fallback = _isolated_ddgs_search()
        if "error" in fallback:
            fallback["attempts_in_process"] = errors[-4:]
        return fallback
    except Exception as e:
        try:
            fallback = _isolated_ddgs_search()
            if "error" in fallback:
                fallback["attempts_in_process"] = errors[-4:] + [str(e)]
            return fallback
        except Exception as sub_e:
            return {
                "error": f"DuckDuckGo 搜尋失敗: {str(e)}; isolated fallback failed: {sub_e}",
                "attempts": errors[-4:],
            }


@tool()
def web_image_search(query: str, max_results: int = 5) -> dict:
    """搜尋網路圖片，回傳圖片標題、來源頁與圖片 URL。"""
    try:
        limit = max(1, min(int(max_results), 20))
        try:
            client = DDGS(verify=False)
        except TypeError:
            client = DDGS()
        rows = list(client.images(query, max_results=limit))
        results = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            results.append({
                "title": row.get("title", ""),
                "image": row.get("image") or row.get("thumbnail") or "",
                "thumbnail": row.get("thumbnail", ""),
                "url": row.get("url", ""),
                "source": row.get("source", ""),
                "width": row.get("width"),
                "height": row.get("height"),
            })
        return {"query": query, "results": results, "count": len(results)}
    except Exception as e:
        return {"error": str(e), "query": query}


@tool()
def download_image(url: str, dest_path: str = "") -> dict:
    """下載圖片到 data/outputs，並回傳可在 WebUI 預覽的 file_url。"""
    try:
        parsed = urllib.parse.urlparse(url)
        suffix = Path(parsed.path).suffix.lower()
        if suffix not in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}:
            suffix = ".jpg"
        default_name = f"image_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}{suffix}"
        out = _resolve_output_path(dest_path, default_name)
        if out.suffix.lower() not in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}:
            out = out.with_suffix(suffix)
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read()
            content_type = resp.headers.get("Content-Type", "")
        if not raw:
            return {"error": "圖片下載內容為空", "url": url}
        out.write_bytes(raw)
        mime = content_type.split(";")[0] if content_type else (mimetypes.guess_type(str(out))[0] or "image/jpeg")
        return {"success": True, "source_url": url, "mime": mime, "size": len(raw), **_file_result(out, mime=mime)}
    except Exception as e:
        return {"error": str(e), "url": url}

@tool()
def ollama_generate(model: str, prompt: str) -> dict:
    """呼叫本地 Ollama 模型推論 (預設本機 11434 埠)"""
    try:
        payload = {"model": model, "prompt": prompt, "stream": False}
        req = urllib.request.Request(
            "http://localhost:11434/api/generate",
            data=json.dumps(payload).encode("utf-8"),
            headers={"Content-Type": "application/json"}
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read().decode())
    except Exception as e:
        return {"error": f"Ollama 本地端未啟動或找不到模型: {str(e)}"}

@tool()
def discord_send(message: str) -> dict:
    """發送訊息至 Discord 頻道"""
    webhook_url = os.environ.get("DISCORD_WEBHOOK_URL")
    if not webhook_url: return {"error": "系統未設定 Discord Webhook"}
    try:
        req = urllib.request.Request(
            webhook_url,
            data=json.dumps({"content": message}).encode("utf-8"),
            headers={"Content-Type": "application/json", "User-Agent": "Agent"}
        )
        with urllib.request.urlopen(req) as resp:
            return {"success": True, "status": resp.status}
    except Exception as e:
        return {"error": str(e)}

@tool()
def telegram_send(message: str) -> dict:
    """透過 Telegram Bot 發送訊息"""
    bot_token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")
    if not bot_token or not chat_id: return {"error": "系統未設定 Telegram 金鑰"}
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    payload = {"chat_id": chat_id, "text": message}
    try:
        req = urllib.request.Request(url, data=json.dumps(payload).encode("utf-8"), headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req) as resp:
            return {"success": True, "body": json.loads(resp.read().decode())}
    except Exception as e:
        return {"error": str(e)}

@tool()
def line_notify(message: str) -> dict:
    """透過 LINE Notify 推送通知"""
    token = os.environ.get("LINE_NOTIFY_TOKEN")
    if not token: return {"error": "系統未設定 LINE Notify Token"}
    try:
        data = urllib.parse.urlencode({"message": message}).encode("utf-8")
        req = urllib.request.Request(
            "https://notify-api.line.me/api/notify",
            data=data,
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/x-www-form-urlencoded"}
        )
        with urllib.request.urlopen(req) as resp:
            return {"success": True, "body": json.loads(resp.read().decode())}
    except Exception as e:
        return {"error": str(e)}

@tool()
def slack_send(message: str) -> dict:
    """發送訊息至 Slack 頻道"""
    webhook_url = os.environ.get("SLACK_WEBHOOK_URL")
    if not webhook_url: return {"error": "系統未設定 Slack Webhook"}
    try:
        req = urllib.request.Request(
            webhook_url,
            data=json.dumps({"text": message}).encode("utf-8"),
            headers={"Content-Type": "application/json"}
        )
        with urllib.request.urlopen(req) as resp:
            return {"success": True, "body": resp.read().decode()}
    except Exception as e:
        return {"error": str(e)}

@tool()
def weather_get(city: str) -> dict:
    """取得指定城市即時天氣（OpenWeatherMap）"""
    api_key = os.environ.get("OPENWEATHER_API_KEY")
    if not api_key: return {"error": "系統未設定天氣 API 金鑰"}
    original = (city or "").strip()
    aliases = {
        "台北": (25.0330, 121.5654, "台北市"),
        "臺北": (25.0330, 121.5654, "台北市"),
        "台北市": (25.0330, 121.5654, "台北市"),
        "臺北市": (25.0330, 121.5654, "台北市"),
        "中正區": (25.0324, 121.5199, "台北市中正區"),
        "台北市中正區": (25.0324, 121.5199, "台北市中正區"),
        "臺北市中正區": (25.0324, 121.5199, "台北市中正區"),
        "新北": (25.0169, 121.4628, "新北市"),
        "新北市": (25.0169, 121.4628, "新北市"),
        "桃園": (24.9937, 121.3009, "桃園市"),
        "桃園市": (24.9937, 121.3009, "桃園市"),
        "台中": (24.1477, 120.6736, "台中市"),
        "臺中": (24.1477, 120.6736, "台中市"),
        "台中市": (24.1477, 120.6736, "台中市"),
        "台南": (22.9999, 120.2270, "台南市"),
        "臺南": (22.9999, 120.2270, "台南市"),
        "高雄": (22.6273, 120.3014, "高雄市"),
        "高雄市": (22.6273, 120.3014, "高雄市"),
    }

    def _get(url: str) -> dict:
        req = urllib.request.Request(url, headers={"User-Agent": "OmniCore-Raphael/1.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode())

    candidates = [original]
    compact = re.sub(r"\s+", "", original)
    if compact and compact not in candidates:
        candidates.append(compact)
    if compact.endswith("區") and "市" in compact:
        city_only = compact.split("市", 1)[0] + "市"
        candidates.append(city_only)
    if compact.endswith("區"):
        candidates.append(compact[:-1])
    candidates.extend(["Taipei,TW"] if "台北" in compact or "臺北" in compact else [])

    errors = []
    for q in [c for i, c in enumerate(candidates) if c and c not in candidates[:i]]:
        url = f"https://api.openweathermap.org/data/2.5/weather?q={urllib.parse.quote(q)}&appid={api_key}&units=metric&lang=zh_tw"
        try:
            data = _get(url)
            data["query_used"] = q
            data["source"] = "openweather:q"
            return data
        except urllib.error.HTTPError as e:
            errors.append(f"{q}: HTTP {e.code}")
        except Exception as e:
            errors.append(f"{q}: {e}")

    coord = aliases.get(compact)
    if coord is None:
        for key, value in aliases.items():
            if key and key in compact:
                coord = value
                break
    if coord is not None:
        lat, lon, label = coord
        url = f"https://api.openweathermap.org/data/2.5/weather?lat={lat}&lon={lon}&appid={api_key}&units=metric&lang=zh_tw"
        try:
            data = _get(url)
            data["query_used"] = label
            data["source"] = "openweather:coordinates"
            return data
        except Exception as e:
            errors.append(f"{label}: {e}")

    return {
        "error": "天氣查詢失敗",
        "city": original,
        "attempts": errors[-6:],
        "suggestion": "請改用城市名稱，例如「台北市」；或補充國家/地區。",
    }

@tool()
def currency_convert(base: str, target: str) -> dict:
    """即時匯率換算（免金鑰開源 API 實作）"""
    url = f"https://open.er-api.com/v6/latest/{base.upper()}"
    try:
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read().decode())
            rates = data.get("rates", {})
            rate = rates.get(target.upper())
            if rate:
                return {"base": base, "target": target, "rate": rate, "time": data.get("time_last_update_utc")}
            return {"error": f"找不到目標貨幣 {target}"}
    except Exception as e:
        return {"error": str(e)}

@tool()
def wikipedia_search(query: str) -> dict:
    """搜尋 Wikipedia 摘要"""
    url = f"https://zh.wikipedia.org/api/rest_v1/page/summary/{urllib.parse.quote(query)}"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "AI-Agent/1.0"})
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read().decode())
            return {"title": data.get("title"), "summary": data.get("extract")}
    except Exception as e:
        return {"error": str(e)}

@tool()
def news_search(keyword: str) -> dict:
    """取得最新新聞頭條（NewsAPI 實作）"""
    api_key = os.environ.get("NEWS_API_KEY")
    if not api_key: return {"error": "系統未設定新聞 API 金鑰"}
    url = f"https://newsapi.org/v2/everything?q={urllib.parse.quote(keyword)}&apiKey={api_key}&pageSize=5&language=zh"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Agent"})
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read().decode())
    except Exception as e:
        return {"error": str(e)}

@tool()
def github_get_repo(repo_name: str) -> dict:
    """取得 GitHub 儲存庫資訊 (格式: owner/repo)"""
    return http_get(f"https://api.github.com/repos/{repo_name}")

@tool()
def github_list_issues(repo_name: str) -> dict:
    """列出 GitHub Issues"""
    return http_get(f"https://api.github.com/repos/{repo_name}/issues")

@tool()
def github_create_issue(repo_name: str, title: str, body: str) -> dict:
    """建立新 GitHub Issue"""
    token = os.environ.get("GITHUB_PAT")
    if not token: return {"error": "系統未設定 GitHub PAT"}
    try:
        req = urllib.request.Request(
            f"https://api.github.com/repos/{repo_name}/issues",
            data=json.dumps({"title": title, "body": body}).encode("utf-8"),
            headers={"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json", "User-Agent": "Agent"}
        )
        with urllib.request.urlopen(req) as resp:
            return {"success": True, "body": json.loads(resp.read().decode())}
    except Exception as e:
        return {"error": str(e)}

@tool()
def pastebin_create(text: str) -> dict:
    """建立 Pastebin 貼文 (使用開源 0x0.st 匿名剪貼簿)"""
    try:
        boundary = "----WebKitFormBoundary7MA4YWxkTrZu0gW"
        data = []
        data.append(f"--{boundary}".encode())
        data.append('Content-Disposition: form-data; name="file"; filename="agent_output.txt"'.encode())
        data.append('Content-Type: text/plain'.encode())
        data.append(''.encode())
        data.append(text.encode())
        data.append(f"--{boundary}--".encode())
        body = b'\r\n'.join(data)

        req = urllib.request.Request(
            "https://0x0.st",
            data=body,
            headers={"Content-Type": f"multipart/form-data; boundary={boundary}", "User-Agent": "curl/7.68.0"},
            method="POST"
        )
        with urllib.request.urlopen(req) as resp:
            return {"url": resp.read().decode().strip()}
    except Exception as e:
        return {"error": str(e)}


# =====================================================================
# 擴充工具庫：檔案搜尋 / 資料處理 / HTTP / Git / 系統診斷
# =====================================================================

def _limit_text(text: str, limit: int = 8000) -> dict:
    text = text or ""
    return {"text": text[:limit], "truncated": len(text) > limit, "length": len(text)}


def _safe_json(value, default=None):
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value)
    except Exception:
        return default


def _run_process(args: list, cwd: str = "", timeout: int = 30) -> dict:
    try:
        res = subprocess.run(
            args,
            cwd=cwd or None,
            capture_output=True,
            text=True,
            timeout=timeout,
            encoding="utf-8",
            errors="replace",
        )
        return {
            "stdout": res.stdout[-8000:],
            "stderr": res.stderr[-4000:],
            "returncode": res.returncode,
            "truncated_stdout": len(res.stdout) > 8000,
            "truncated_stderr": len(res.stderr) > 4000,
        }
    except subprocess.TimeoutExpired:
        return {"error": f"執行逾時（{timeout} 秒）", "args": args}
    except Exception as e:
        return {"error": str(e), "args": args}


def _file_result(path: Path, **extra) -> dict:
    path = path.resolve()
    out = {"path": str(path), **extra}
    try:
        rel = path.relative_to(DATA_DIR.resolve())
        out["file_url"] = "/files/" + "/".join(rel.parts)
        out["filename"] = path.name
    except Exception:
        pass
    return out


def _pyautogui():
    try:
        import pyautogui
        pyautogui.PAUSE = 0.05
        pyautogui.FAILSAFE = True
        return pyautogui, None
    except Exception as e:
        return None, f"pyautogui 未安裝或無法使用：{e}"


def _window_error() -> str | None:
    if platform.system().lower() != "windows":
        return "視窗工具目前只支援 Windows。"
    if not getattr(ctypes, "windll", None):
        return "目前環境無法存取 Windows 視窗 API。"
    return None


def _int_handle(value) -> int:
    try:
        raw = value.value if hasattr(value, "value") else value
        return int(raw or 0)
    except Exception:
        return 0


def _process_name(pid: int) -> str:
    try:
        import psutil
        return psutil.Process(int(pid)).name()
    except Exception:
        return ""


def _window_rect(hwnd: int) -> dict:
    rect = wintypes.RECT()
    ok = ctypes.windll.user32.GetWindowRect(int(hwnd), ctypes.byref(rect))
    if not ok:
        return {}
    left, top, right, bottom = int(rect.left), int(rect.top), int(rect.right), int(rect.bottom)
    return {
        "left": left,
        "top": top,
        "right": right,
        "bottom": bottom,
        "width": max(0, right - left),
        "height": max(0, bottom - top),
    }


def _window_info(hwnd: int) -> dict:
    err = _window_error()
    if err or not hwnd:
        return {}
    user32 = ctypes.windll.user32
    hwnd = int(hwnd)
    length = user32.GetWindowTextLengthW(hwnd)
    title_buf = ctypes.create_unicode_buffer(max(1, int(length) + 1))
    user32.GetWindowTextW(hwnd, title_buf, len(title_buf))
    pid = wintypes.DWORD()
    user32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
    rect = _window_rect(hwnd)
    return {
        "hwnd": str(hwnd),
        "title": title_buf.value,
        "process": _process_name(int(pid.value)),
        "pid": int(pid.value),
        "rect": rect,
        "active": hwnd == int(user32.GetForegroundWindow() or 0),
        "visible": bool(user32.IsWindowVisible(hwnd)),
        "minimized": bool(user32.IsIconic(hwnd)),
    }


def _active_window_info() -> dict:
    err = _window_error()
    if err:
        return {"error": err}
    hwnd = ctypes.windll.user32.GetForegroundWindow()
    info = _window_info(_int_handle(hwnd))
    return info or {"error": "找不到目前前景視窗。"}


def _visible_windows(limit: int = 80) -> list[dict]:
    err = _window_error()
    if err:
        return []
    user32 = ctypes.windll.user32
    windows: list[dict] = []

    @ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
    def enum_proc(hwnd, lparam):
        hwnd_i = _int_handle(hwnd)
        if not hwnd_i or not user32.IsWindowVisible(hwnd_i):
            return True
        info = _window_info(hwnd_i)
        title = (info.get("title") or "").strip()
        rect = info.get("rect") or {}
        if title and rect.get("width", 0) > 0 and rect.get("height", 0) > 0:
            windows.append(info)
        return True

    user32.EnumWindows(enum_proc, 0)
    windows.sort(key=lambda w: (not w.get("active"), w.get("title", "").lower()))
    return windows[:max(1, min(int(limit or 80), 300))]


def _match_window(target: str) -> tuple[dict | None, list[dict], str]:
    target = str(target or "").strip()
    if not target:
        active = _active_window_info()
        if active.get("error"):
            return None, [], active["error"]
        return active, [active], ""
    windows = _visible_windows(200)
    target_l = target.lower()
    exact_hwnd = None
    try:
        exact_hwnd = int(target, 0)
    except Exception:
        exact_hwnd = None
    matches = []
    for win in windows:
        if exact_hwnd is not None and str(win.get("hwnd")) == str(exact_hwnd):
            matches = [win]
            break
        haystack = f"{win.get('title', '')} {win.get('process', '')} {win.get('pid', '')}".lower()
        if target_l in haystack:
            matches.append(win)
    if not matches:
        return None, [], f"找不到符合「{target}」的可見視窗。"
    return matches[0], matches, ""


@tool()
def computer_active_window() -> dict:
    """取得目前前景視窗，供桌面操作前確認目標。"""
    info = _active_window_info()
    if info.get("error"):
        return {"error": info["error"]}
    return {"success": True, "active_window": info}


@tool()
def computer_list_windows(max_items: int = 40) -> dict:
    """列出目前可見視窗，包含標題、程序、PID、位置與 hwnd。"""
    err = _window_error()
    if err:
        return {"error": err}
    windows = _visible_windows(max_items)
    return {
        "success": True,
        "active_window": _active_window_info(),
        "windows": windows,
        "count": len(windows),
    }


@tool()
def computer_focus_window(target: str) -> dict:
    """依視窗標題、程序名、PID 或 hwnd 找到並切換到指定視窗。"""
    err = _window_error()
    if err:
        return {"error": err}
    selected, matches, message = _match_window(target)
    if not selected:
        return {"error": message, "target": target, "active_window": _active_window_info()}
    user32 = ctypes.windll.user32
    hwnd = int(selected["hwnd"])
    try:
        user32.ShowWindow(hwnd, 9)
        ok = bool(user32.SetForegroundWindow(hwnd))
    except Exception as e:
        return {"error": str(e), "target": target, "selected_window": selected}
    return {
        "success": True,
        "focused": ok,
        "selected_window": selected,
        "active_window": _active_window_info(),
        "match_count": len(matches),
        "matches": matches[:8],
    }


@tool()
def computer_screenshot_window(target: str = "", path: str = "") -> dict:
    """只截取目前前景視窗，或符合 target 的指定視窗；target 可為標題、程序、PID 或 hwnd。"""
    selected, matches, message = _match_window(target)
    if not selected:
        return {"error": message, "target": target, "active_window": _active_window_info()}
    if selected.get("minimized"):
        return {"error": "目標視窗目前最小化，請先切換或還原視窗。", "selected_window": selected}
    rect = selected.get("rect") or {}
    if rect.get("width", 0) <= 0 or rect.get("height", 0) <= 0:
        return {"error": "目標視窗大小無效，無法截圖。", "selected_window": selected}
    try:
        out = _resolve_output_path(path, f"window_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
        if out.suffix.lower() != ".png":
            out = out.with_suffix(".png")
        bbox = (rect["left"], rect["top"], rect["right"], rect["bottom"])
        from PIL import ImageGrab
        img = ImageGrab.grab(bbox=bbox)
        img.save(out)
        return {
            "success": True,
            **_file_result(out),
            "width": img.width,
            "height": img.height,
            "window": selected,
            "match_count": len(matches),
        }
    except Exception as e:
        return {"error": str(e), "selected_window": selected}


@tool()
def path_exists(path: str) -> dict:
    """檢查檔案或資料夾是否存在，並回傳基本類型。"""
    p = Path(path)
    return {"exists": p.exists(), "is_file": p.is_file(), "is_dir": p.is_dir(), "path": str(p)}


@tool()
def make_directory(path: str) -> dict:
    """建立資料夾；父層不存在時會一併建立。"""
    try:
        p = Path(path)
        p.mkdir(parents=True, exist_ok=True)
        return {"success": True, "path": str(p.resolve())}
    except Exception as e:
        return {"error": str(e)}


@tool()
def read_file_range(path: str, start_line: int = 1, end_line: int = 120) -> dict:
    """讀取檔案指定行數範圍，適合檢查大型文字檔。"""
    try:
        p = Path(path)
        if not p.exists():
            return {"error": f"檔案不存在：{path}"}
        start_line = max(1, int(start_line))
        end_line = max(start_line, int(end_line))
        rows = []
        with p.open("r", encoding="utf-8", errors="replace") as f:
            for idx, line in enumerate(f, start=1):
                if idx < start_line:
                    continue
                if idx > end_line:
                    break
                rows.append({"line": idx, "text": line.rstrip("\n")})
        return {"path": str(p), "start_line": start_line, "end_line": end_line, "lines": rows}
    except Exception as e:
        return {"error": str(e)}


@tool()
def list_directory_recursive(path: str, pattern: str = "*", max_items: int = 200) -> dict:
    """遞迴列出資料夾內容，可用 glob pattern 篩選。"""
    try:
        root = Path(path)
        if not root.exists():
            return {"error": f"路徑不存在：{path}"}
        max_items = max(1, min(int(max_items), 2000))
        items = []
        for p in root.rglob(pattern or "*"):
            try:
                st = p.stat()
                items.append({
                    "path": str(p),
                    "relative": str(p.relative_to(root)),
                    "is_file": p.is_file(),
                    "is_dir": p.is_dir(),
                    "size": st.st_size if p.is_file() else 0,
                    "modified": datetime.datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds"),
                })
            except Exception:
                continue
            if len(items) >= max_items:
                break
        return {"root": str(root), "pattern": pattern, "items": items, "count": len(items), "truncated": len(items) >= max_items}
    except Exception as e:
        return {"error": str(e)}


@tool()
def search_files(root: str, pattern: str, max_results: int = 100) -> dict:
    """依檔名 glob pattern 搜尋檔案，例如 *.py 或 *report*.pdf。"""
    try:
        base = Path(root)
        max_results = max(1, min(int(max_results), 1000))
        matches = []
        for p in base.rglob(pattern):
            matches.append({"path": str(p), "is_file": p.is_file(), "size": p.stat().st_size if p.is_file() else 0})
            if len(matches) >= max_results:
                break
        return {"root": str(base), "pattern": pattern, "results": matches, "count": len(matches), "truncated": len(matches) >= max_results}
    except Exception as e:
        return {"error": str(e)}


@tool()
def search_text(root: str, query: str, file_pattern: str = "*", max_results: int = 80) -> dict:
    """在文字檔中搜尋關鍵字或正規表示式，回傳檔案、行號與片段。"""
    try:
        base = Path(root)
        max_results = max(1, min(int(max_results), 500))
        rx = re.compile(query, re.IGNORECASE)
        hits = []
        for p in base.rglob(file_pattern or "*"):
            if not p.is_file() or p.stat().st_size > 3_000_000:
                continue
            try:
                with p.open("r", encoding="utf-8", errors="ignore") as f:
                    for line_no, line in enumerate(f, start=1):
                        if rx.search(line):
                            hits.append({"path": str(p), "line": line_no, "text": line.strip()[:500]})
                            if len(hits) >= max_results:
                                return {"root": str(base), "query": query, "results": hits, "count": len(hits), "truncated": True}
            except Exception:
                continue
        return {"root": str(base), "query": query, "results": hits, "count": len(hits), "truncated": False}
    except re.error as e:
        return {"error": f"正規表示式錯誤：{e}"}
    except Exception as e:
        return {"error": str(e)}


@tool()
def replace_in_file(path: str, old: str, new: str, count: int = 0) -> dict:
    """在文字檔內取代字串；count=0 代表全部取代。"""
    try:
        p = Path(path)
        text = p.read_text(encoding="utf-8", errors="replace")
        replaced = text.count(old) if count == 0 else min(text.count(old), int(count))
        if replaced == 0:
            return {"success": False, "message": "找不到要取代的文字", "path": str(p)}
        out = text.replace(old, new, int(count) if count else -1)
        p.write_text(out, encoding="utf-8")
        return {"success": True, "path": str(p.resolve()), "replaced": replaced}
    except Exception as e:
        return {"error": str(e)}


@tool(schema_override={"properties": {"algorithm": {"enum": ["md5", "sha1", "sha256", "sha512"]}}})
def file_hash(path: str, algorithm: str = "sha256") -> dict:
    """計算檔案雜湊值，用於比對檔案完整性。"""
    try:
        h = hashlib.new(algorithm)
        p = Path(path)
        with p.open("rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                h.update(chunk)
        return {"path": str(p), "algorithm": algorithm, "hash": h.hexdigest(), "size": p.stat().st_size}
    except Exception as e:
        return {"error": str(e)}


@tool()
def detect_file_type(path: str) -> dict:
    """依副檔名與少量檔頭資訊推測檔案類型。"""
    try:
        p = Path(path)
        mime, encoding = mimetypes.guess_type(str(p))
        head = p.read_bytes()[:16] if p.is_file() else b""
        return {"path": str(p), "mime": mime, "encoding": encoding, "head_hex": head.hex(), "suffix": p.suffix}
    except Exception as e:
        return {"error": str(e)}


@tool()
def zip_create(zip_path: str, paths: list) -> dict:
    """建立 zip 壓縮檔；paths 為檔案或資料夾路徑清單。"""
    try:
        zp = Path(zip_path)
        zp.parent.mkdir(parents=True, exist_ok=True)
        added = 0
        with zipfile.ZipFile(zp, "w", zipfile.ZIP_DEFLATED) as zf:
            for raw in paths:
                p = Path(str(raw))
                if p.is_dir():
                    for child in p.rglob("*"):
                        if child.is_file():
                            zf.write(child, child.relative_to(p.parent))
                            added += 1
                elif p.is_file():
                    zf.write(p, p.name)
                    added += 1
        return {"success": True, "zip_path": str(zp.resolve()), "files_added": added}
    except Exception as e:
        return {"error": str(e)}


@tool()
def zip_extract(zip_path: str, dest_dir: str) -> dict:
    """解壓縮 zip 檔到指定資料夾。"""
    try:
        zp = Path(zip_path)
        dest = Path(dest_dir)
        dest.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(zp) as zf:
            names = zf.namelist()
            zf.extractall(dest)
        return {"success": True, "dest_dir": str(dest.resolve()), "files": names[:200], "count": len(names)}
    except Exception as e:
        return {"error": str(e)}


@tool()
def download_file(url: str, dest_path: str) -> dict:
    """下載 URL 內容到本地檔案。"""
    try:
        dest = Path(dest_path)
        dest.parent.mkdir(parents=True, exist_ok=True)
        req = urllib.request.Request(url, headers={"User-Agent": "OmniCore-Raphael/1.0"})
        with urllib.request.urlopen(req, timeout=30) as resp, dest.open("wb") as f:
            shutil.copyfileobj(resp, f)
            status = getattr(resp, "status", None)
        return {"success": True, "url": url, "path": str(dest.resolve()), "size": dest.stat().st_size, "status": status}
    except Exception as e:
        return {"error": str(e), "url": url}


@tool()
def http_request(method: str, url: str, headers: dict = None, body: str = "", timeout: int = 20) -> dict:
    """發送 HTTP 請求，支援 GET/POST/PUT/PATCH/DELETE 與自訂 headers/body。"""
    try:
        method = (method or "GET").upper()
        if method not in {"GET", "POST", "PUT", "PATCH", "DELETE", "HEAD"}:
            return {"error": f"不支援的 HTTP 方法：{method}"}
        data = body.encode("utf-8") if body and method not in {"GET", "HEAD"} else None
        req_headers = {"User-Agent": "OmniCore-Raphael/1.0"}
        if isinstance(headers, dict):
            req_headers.update({str(k): str(v) for k, v in headers.items()})
        req = urllib.request.Request(url, data=data, headers=req_headers, method=method)
        with urllib.request.urlopen(req, timeout=max(1, int(timeout))) as resp:
            raw = resp.read()
            text = raw.decode("utf-8", errors="replace")
            return {"status": resp.status, "headers": dict(resp.headers), **_limit_text(text, 6000)}
    except urllib.error.HTTPError as e:
        text = e.read().decode("utf-8", errors="replace")
        return {"error": f"HTTP {e.code}", "status": e.code, **_limit_text(text, 3000)}
    except Exception as e:
        return {"error": str(e), "url": url}


@tool()
def json_parse(text: str) -> dict:
    """解析 JSON 字串並回傳簡要結構。"""
    try:
        obj = json.loads(text)
        if isinstance(obj, dict):
            return {"type": "object", "keys": list(obj.keys())[:100], "value": obj}
        if isinstance(obj, list):
            return {"type": "array", "length": len(obj), "sample": obj[:5]}
        return {"type": type(obj).__name__, "value": obj}
    except Exception as e:
        return {"error": str(e)}


@tool()
def json_query(text: str, path: str) -> dict:
    """用簡單 dot path 查詢 JSON，例如 data.items.0.title。"""
    try:
        cur = json.loads(text)
        for part in [p for p in path.split(".") if p != ""]:
            if isinstance(cur, list):
                cur = cur[int(part)]
            elif isinstance(cur, dict):
                cur = cur[part]
            else:
                return {"error": f"無法在 {type(cur).__name__} 上繼續查詢 {part}"}
        return {"path": path, "value": cur}
    except Exception as e:
        return {"error": str(e), "path": path}


@tool()
def csv_read(path: str, max_rows: int = 50) -> dict:
    """讀取 CSV 檔前 N 列。"""
    try:
        rows = []
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f)
            for idx, row in enumerate(reader):
                if idx >= max(1, min(int(max_rows), 1000)):
                    break
                rows.append(dict(row))
        return {"path": path, "columns": reader.fieldnames or [], "rows": rows, "count": len(rows)}
    except Exception as e:
        return {"error": str(e)}


@tool()
def csv_write(path: str, rows: list) -> dict:
    """將物件陣列寫成 CSV。rows 需為 [{欄位: 值}] 格式。"""
    try:
        if not rows:
            return {"error": "rows 不可為空"}
        rows = [r for r in rows if isinstance(r, dict)]
        fields = list(dict.fromkeys(k for row in rows for k in row.keys()))
        p = Path(path)
        p.parent.mkdir(parents=True, exist_ok=True)
        with p.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            writer.writerows(rows)
        return {"success": True, "path": str(p.resolve()), "rows": len(rows), "columns": fields}
    except Exception as e:
        return {"error": str(e)}


@tool()
def sqlite_query(db_path: str, query: str, params: list = None, max_rows: int = 100) -> dict:
    """執行 SQLite 查詢；SELECT 會回傳列資料，其他語句會提交變更。"""
    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cur = conn.execute(query, params or [])
        if query.strip().lower().startswith(("select", "with", "pragma")):
            rows = [dict(r) for r in cur.fetchmany(max(1, min(int(max_rows), 1000)))]
            conn.close()
            return {"rows": rows, "count": len(rows)}
        conn.commit()
        changed = cur.rowcount
        conn.close()
        return {"success": True, "changed": changed}
    except Exception as e:
        return {"error": str(e)}


@tool()
def sqlite_schema(db_path: str) -> dict:
    """列出 SQLite 資料庫表格與欄位。"""
    try:
        conn = sqlite3.connect(db_path)
        tables = [r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")]
        out = {}
        for table in tables:
            out[table] = [dict(cid=r[0], name=r[1], type=r[2], notnull=r[3], default=r[4], pk=r[5]) for r in conn.execute(f"PRAGMA table_info({table})")]
        conn.close()
        return {"tables": out}
    except Exception as e:
        return {"error": str(e)}


@tool()
def git_status(repo_path: str = ".") -> dict:
    """取得 Git 工作區狀態。"""
    return _run_process(["git", "status", "--short"], cwd=repo_path, timeout=20)


@tool()
def git_log(repo_path: str = ".", max_count: int = 5) -> dict:
    """取得 Git 最近提交紀錄。"""
    return _run_process(["git", "log", f"--max-count={max(1, min(int(max_count), 50))}", "--oneline", "--decorate"], cwd=repo_path, timeout=20)


@tool()
def git_diff(repo_path: str = ".", path: str = "") -> dict:
    """取得 Git diff，path 留空代表全部變更。"""
    args = ["git", "diff", "--"]
    if path:
        args.append(path)
    return _run_process(args, cwd=repo_path, timeout=30)


@tool()
def git_show(repo_path: str, revision: str = "HEAD", path: str = "") -> dict:
    """查看 Git 指定 revision 或檔案內容。"""
    args = ["git", "show", revision]
    if path:
        args.extend(["--", path])
    return _run_process(args, cwd=repo_path, timeout=30)


@tool()
def git_branch(repo_path: str = ".") -> dict:
    """列出 Git 分支並顯示目前分支。"""
    return _run_process(["git", "branch", "--show-current"], cwd=repo_path, timeout=10) | {
        "branches": _run_process(["git", "branch", "--list"], cwd=repo_path, timeout=10).get("stdout", "")
    }


@tool()
def system_info() -> dict:
    """取得作業系統、Python、CPU 與記憶體基本資訊。"""
    info = {
        "platform": platform.platform(),
        "python": sys.version,
        "machine": platform.machine(),
        "processor": platform.processor(),
        "cwd": os.getcwd(),
    }
    try:
        import psutil
        info["cpu_count"] = psutil.cpu_count()
        info["memory"] = dict(psutil.virtual_memory()._asdict())
    except Exception:
        pass
    return info


@tool()
def disk_usage(path: str = ".") -> dict:
    """取得指定磁碟/路徑的容量使用狀況。"""
    try:
        total, used, free = shutil.disk_usage(path)
        return {"path": path, "total": total, "used": used, "free": free, "used_percent": round(used / total * 100, 2)}
    except Exception as e:
        return {"error": str(e)}


@tool()
def network_ping(host: str, count: int = 4) -> dict:
    """Ping 主機並回傳結果摘要。"""
    count = max(1, min(int(count), 10))
    flag = "-n" if os.name == "nt" else "-c"
    return _run_process(["ping", flag, str(count), host], timeout=20)


@tool()
def dns_lookup(host: str) -> dict:
    """查詢主機 DNS 解析結果。"""
    try:
        return {"host": host, "addresses": sorted({r[4][0] for r in socket.getaddrinfo(host, None)})}
    except Exception as e:
        return {"error": str(e), "host": host}


@tool()
def port_check(host: str, port: int, timeout: float = 3.0) -> dict:
    """檢查 TCP 主機連接埠是否可連線。"""
    try:
        started = time.perf_counter()
        with socket.create_connection((host, int(port)), timeout=max(0.1, float(timeout))):
            pass
        return {"host": host, "port": int(port), "open": True, "latency_ms": round((time.perf_counter() - started) * 1000, 2)}
    except Exception as e:
        return {"host": host, "port": int(port), "open": False, "error": str(e)}


@tool()
def python_run(code: str, timeout: int = 10) -> dict:
    """在隔離 Python 子程序執行短程式，適合資料轉換與快速計算。"""
    return _run_process([sys.executable, "-c", code], timeout=max(1, min(int(timeout), 60)))


@tool()
def regex_extract(text: str, pattern: str, max_results: int = 50) -> dict:
    """用正規表示式從文字中擷取匹配結果。"""
    try:
        rx = re.compile(pattern, re.MULTILINE)
        out = []
        for m in rx.finditer(text):
            out.append({"match": m.group(0), "groups": list(m.groups()), "span": list(m.span())})
            if len(out) >= max(1, min(int(max_results), 500)):
                break
        return {"results": out, "count": len(out)}
    except Exception as e:
        return {"error": str(e)}


@tool()
def text_summarize_basic(text: str, max_sentences: int = 5) -> dict:
    """用簡單規則抽取文字前幾個重點句，無需外部模型。"""
    parts = re.split(r"(?<=[。！？.!?])\s+|\n+", text.strip())
    parts = [p.strip() for p in parts if p.strip()]
    return {"summary": parts[:max(1, min(int(max_sentences), 20))], "sentence_count": len(parts)}


@tool()
def computer_screenshot(path: str = "", region: list = None) -> dict:
    """擷取目前螢幕畫面並存成 PNG；region 可傳 [x, y, width, height]。"""
    try:
        out = _resolve_output_path(path, f"screenshot_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
        if out.suffix.lower() != ".png":
            out = out.with_suffix(".png")
        pyautogui, err = _pyautogui()
        if pyautogui:
            img = pyautogui.screenshot(region=tuple(region) if region else None)
            img.save(out)
        else:
            from PIL import ImageGrab
            bbox = None
            if region and len(region) == 4:
                x, y, w, h = [int(v) for v in region]
                bbox = (x, y, x + w, y + h)
            img = ImageGrab.grab(bbox=bbox)
            img.save(out)
        return {
            "success": True,
            **_file_result(out),
            "width": img.width,
            "height": img.height,
            "active_window": _active_window_info(),
            "region": region or None,
        }
    except Exception as e:
        return {"error": str(e)}


@tool()
def computer_screen_size() -> dict:
    """取得目前主螢幕尺寸。"""
    pyautogui, err = _pyautogui()
    if not pyautogui:
        return {"error": err}
    try:
        size = pyautogui.size()
        return {"width": size.width, "height": size.height}
    except Exception as e:
        return {"error": str(e)}


@tool()
def computer_mouse_position() -> dict:
    """取得目前滑鼠座標。"""
    pyautogui, err = _pyautogui()
    if not pyautogui:
        return {"error": err}
    try:
        p = pyautogui.position()
        return {"x": p.x, "y": p.y}
    except Exception as e:
        return {"error": str(e)}


@tool()
def computer_click(x: int, y: int, button: str = "left", clicks: int = 1, interval: float = 0.05) -> dict:
    """在指定螢幕座標點擊滑鼠。button 可為 left/right/middle。"""
    pyautogui, err = _pyautogui()
    if not pyautogui:
        return {"error": err}
    try:
        before = _active_window_info()
        pyautogui.click(int(x), int(y), clicks=max(1, int(clicks)), interval=max(0.0, float(interval)), button=button)
        return {
            "success": True,
            "x": int(x),
            "y": int(y),
            "button": button,
            "clicks": clicks,
            "active_window_before": before,
            "active_window_after": _active_window_info(),
        }
    except Exception as e:
        return {"error": str(e)}


@tool()
def computer_double_click(x: int, y: int, button: str = "left") -> dict:
    """在指定螢幕座標雙擊滑鼠。"""
    return computer_click(x, y, button=button, clicks=2, interval=0.06)


@tool()
def computer_move_mouse(x: int, y: int, duration: float = 0.1) -> dict:
    """將滑鼠移動到指定座標。"""
    pyautogui, err = _pyautogui()
    if not pyautogui:
        return {"error": err}
    try:
        pyautogui.moveTo(int(x), int(y), duration=max(0.0, float(duration)))
        return {"success": True, "x": int(x), "y": int(y)}
    except Exception as e:
        return {"error": str(e)}


@tool()
def computer_drag_mouse(x: int, y: int, duration: float = 0.3, button: str = "left") -> dict:
    """按住滑鼠拖曳到指定座標。"""
    pyautogui, err = _pyautogui()
    if not pyautogui:
        return {"error": err}
    try:
        pyautogui.dragTo(int(x), int(y), duration=max(0.0, float(duration)), button=button)
        return {"success": True, "x": int(x), "y": int(y), "button": button}
    except Exception as e:
        return {"error": str(e)}


@tool()
def computer_scroll(amount: int, x: int = 0, y: int = 0) -> dict:
    """滾動滑鼠滾輪；正數向上、負數向下。若 x/y 非 0 會先移到該座標。"""
    pyautogui, err = _pyautogui()
    if not pyautogui:
        return {"error": err}
    try:
        if x or y:
            pyautogui.moveTo(int(x), int(y), duration=0.05)
        pyautogui.scroll(int(amount))
        return {"success": True, "amount": int(amount)}
    except Exception as e:
        return {"error": str(e)}


@tool()
def computer_type_text(text: str, interval: float = 0.01, paste_for_unicode: bool = True) -> dict:
    """模擬真實鍵盤輸入文字；中文等 Unicode 預設用剪貼簿貼上以確保成功。"""
    pyautogui, err = _pyautogui()
    if not pyautogui:
        return {"error": err}
    try:
        active = _active_window_info()
        ascii_only = all(ord(ch) < 128 for ch in text)
        if paste_for_unicode and not ascii_only:
            try:
                import pyperclip
                pyperclip.copy(text)
                pyautogui.hotkey("ctrl", "v")
                return {"success": True, "method": "clipboard_paste", "chars": len(text), "active_window": active}
            except Exception:
                pass
        pyautogui.write(text, interval=max(0.0, float(interval)))
        return {"success": True, "method": "key_write", "chars": len(text), "active_window": active}
    except Exception as e:
        return {"error": str(e)}


@tool()
def computer_press_key(key: str, presses: int = 1, interval: float = 0.05) -> dict:
    """按下單一鍵，例如 enter、esc、tab、backspace、space。"""
    pyautogui, err = _pyautogui()
    if not pyautogui:
        return {"error": err}
    try:
        pyautogui.press(key, presses=max(1, int(presses)), interval=max(0.0, float(interval)))
        return {"success": True, "key": key, "presses": presses}
    except Exception as e:
        return {"error": str(e)}


@tool()
def computer_hotkey(keys: list) -> dict:
    """按下快捷鍵組合，例如 ['ctrl','c']、['alt','tab']、['win','r']。"""
    pyautogui, err = _pyautogui()
    if not pyautogui:
        return {"error": err}
    try:
        pyautogui.hotkey(*[str(k) for k in keys])
        return {"success": True, "keys": keys}
    except Exception as e:
        return {"error": str(e)}


@tool()
def computer_locate_image(image_path: str, confidence: float = 0.8) -> dict:
    """在螢幕中尋找指定圖片位置；需要 pyautogui，confidence 可能需要 opencv。"""
    pyautogui, err = _pyautogui()
    if not pyautogui:
        return {"error": err}
    try:
        box = pyautogui.locateOnScreen(image_path, confidence=float(confidence))
        if not box:
            return {"found": False, "image_path": image_path}
        center = pyautogui.center(box)
        return {"found": True, "box": {"left": box.left, "top": box.top, "width": box.width, "height": box.height}, "center": {"x": center.x, "y": center.y}}
    except Exception as e:
        return {"error": str(e), "image_path": image_path}


@tool()
def computer_control(steps: list, screenshot_after: bool = True) -> dict:
    """依序執行電腦操作步驟。每步格式：{action:'focus_window/click/type/hotkey/press/scroll/move/drag/screenshot/screenshot_window', ...}。"""
    steps = steps or []
    unsafe_actions = {"click", "double_click", "type", "hotkey", "press", "scroll", "move", "drag"}
    awareness_actions = {"active_window", "list_windows", "focus_window", "screenshot_window"}
    normalized_actions = [
        str(step.get("action", "")).lower()
        for step in steps
        if isinstance(step, dict)
    ]
    if any(action in unsafe_actions for action in normalized_actions):
        first_action = normalized_actions[0] if normalized_actions else ""
        if first_action not in awareness_actions:
            return {
                "error": "拒絕執行批次電腦操作：第一步必須先確認或切換目標視窗。",
                "required_first_actions": sorted(awareness_actions),
                "recovery_hint": "請先用 active_window/list_windows/focus_window/screenshot_window 確認目標視窗，再執行點擊或輸入。",
            }
    results = []
    for i, step in enumerate(steps, start=1):
        if not isinstance(step, dict):
            results.append({"step": i, "error": "步驟必須是物件"})
            continue
        action = str(step.get("action", "")).lower()
        if action == "focus_window":
            res = computer_focus_window(step.get("target", ""))
        elif action == "active_window":
            res = computer_active_window()
        elif action == "list_windows":
            res = computer_list_windows(step.get("max_items", 40))
        elif action == "click":
            res = computer_click(step.get("x", 0), step.get("y", 0), step.get("button", "left"), step.get("clicks", 1), step.get("interval", 0.05))
        elif action == "double_click":
            res = computer_double_click(step.get("x", 0), step.get("y", 0), step.get("button", "left"))
        elif action == "type":
            res = computer_type_text(step.get("text", ""), step.get("interval", 0.01), step.get("paste_for_unicode", True))
        elif action == "hotkey":
            res = computer_hotkey(step.get("keys", []))
        elif action == "press":
            res = computer_press_key(step.get("key", ""), step.get("presses", 1), step.get("interval", 0.05))
        elif action == "scroll":
            res = computer_scroll(step.get("amount", 0), step.get("x", 0), step.get("y", 0))
        elif action == "move":
            res = computer_move_mouse(step.get("x", 0), step.get("y", 0), step.get("duration", 0.1))
        elif action == "drag":
            res = computer_drag_mouse(step.get("x", 0), step.get("y", 0), step.get("duration", 0.3), step.get("button", "left"))
        elif action == "screenshot":
            res = computer_screenshot(step.get("path", ""), step.get("region"))
        elif action == "screenshot_window":
            res = computer_screenshot_window(step.get("target", ""), step.get("path", ""))
        else:
            res = {"error": f"未知 action：{action}"}
        results.append({"step": i, "action": action, "result": res})
        delay = float(step.get("delay", 0) or 0)
        if delay > 0:
            time.sleep(min(delay, 10))
    final_shot = computer_screenshot_window() if screenshot_after else None
    return {"success": all("error" not in r.get("result", {}) for r in results), "steps": results, "screenshot": final_shot}


@tool()
def base64_encode_text(text: str) -> dict:
    """將文字編碼為 Base64。"""
    try:
        return {"base64": base64.b64encode(text.encode("utf-8")).decode("ascii")}
    except Exception as e:
        return {"error": str(e)}


@tool()
def base64_decode_text(data: str) -> dict:
    """將 Base64 解碼為 UTF-8 文字。"""
    try:
        raw = base64.b64decode(data)
        return {"text": raw.decode("utf-8", errors="replace")}
    except Exception as e:
        return {"error": str(e)}


@tool()
def url_encode(text: str) -> dict:
    """URL encode 文字。"""
    return {"encoded": urllib.parse.quote(text)}


@tool()
def url_decode(text: str) -> dict:
    """URL decode 文字。"""
    return {"decoded": urllib.parse.unquote(text)}


@tool()
def uuid_generate(count: int = 1) -> dict:
    """產生 UUID v4。"""
    count = max(1, min(int(count), 100))
    return {"uuids": [str(uuid.uuid4()) for _ in range(count)]}


@tool()
def timestamp_convert(timestamp: float, timezone_offset_hours: float = 8) -> dict:
    """將 Unix timestamp 轉換為日期時間。"""
    try:
        tz = datetime.timezone(datetime.timedelta(hours=float(timezone_offset_hours)))
        dt = datetime.datetime.fromtimestamp(float(timestamp), tz=tz)
        return {"timestamp": timestamp, "datetime": dt.isoformat(), "timezone_offset_hours": timezone_offset_hours}
    except Exception as e:
        return {"error": str(e)}


@tool()
def html_extract_text(html: str) -> dict:
    """從 HTML 字串中粗略抽出純文字內容。"""
    try:
        text = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", html)
        text = re.sub(r"(?s)<[^>]+>", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        return _limit_text(text, 8000)
    except Exception as e:
        return {"error": str(e)}


@tool()
def image_info(path: str) -> dict:
    """讀取圖片尺寸、格式與基本 metadata。"""
    try:
        from PIL import Image
        p = Path(path)
        with Image.open(p) as img:
            return {
                "path": str(p),
                "format": img.format,
                "mode": img.mode,
                "width": img.width,
                "height": img.height,
                "size": p.stat().st_size,
            }
    except Exception as e:
        return {"error": str(e)}


@tool()
def image_resize(src_path: str, dest_path: str, width: int, height: int = 0, keep_aspect: bool = True) -> dict:
    """調整圖片尺寸並輸出到新檔案；height=0 且 keep_aspect=true 時自動等比例。"""
    try:
        from PIL import Image
        src = Path(src_path)
        dest = Path(dest_path)
        dest.parent.mkdir(parents=True, exist_ok=True)
        with Image.open(src) as img:
            width = max(1, int(width))
            height = int(height)
            if keep_aspect:
                ratio = width / img.width
                height = max(1, int(img.height * ratio)) if height <= 0 else height
                img = img.resize((width, height), Image.LANCZOS)
            else:
                img = img.resize((width, max(1, height)), Image.LANCZOS)
            img.save(dest)
        return {"success": True, "src": str(src), "dest": str(dest.resolve()), "width": width, "height": height}
    except Exception as e:
        return {"error": str(e)}


@tool()
def pdf_extract_text(path: str, max_pages: int = 10) -> dict:
    """從 PDF 擷取文字；需要 pypdf 或 PyPDF2。"""
    try:
        try:
            from pypdf import PdfReader
        except Exception:
            from PyPDF2 import PdfReader
        reader = PdfReader(path)
        pages = []
        for i, page in enumerate(reader.pages[:max(1, min(int(max_pages), 100))], start=1):
            pages.append({"page": i, "text": (page.extract_text() or "")[:5000]})
        return {"path": path, "page_count": len(reader.pages), "pages": pages}
    except Exception as e:
        return {"error": str(e)}


@tool()
def docx_extract_text(path: str) -> dict:
    """從 Word docx 擷取段落文字；需要 python-docx。"""
    try:
        import docx
        document = docx.Document(path)
        paragraphs = [p.text for p in document.paragraphs if p.text.strip()]
        return {"path": path, "paragraphs": paragraphs[:500], "count": len(paragraphs)}
    except Exception as e:
        return {"error": str(e)}


@tool()
def xlsx_read_sheet(path: str, sheet_name: str = "", max_rows: int = 50) -> dict:
    """讀取 Excel xlsx 工作表前 N 列；需要 openpyxl。"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx > max(1, min(int(max_rows), 1000)):
                break
            rows.append(list(row))
        return {"path": path, "sheet": ws.title, "rows": rows, "count": len(rows), "sheets": wb.sheetnames}
    except Exception as e:
        return {"error": str(e)}


@tool()
def xlsx_write_sheet(path: str, sheet_name: str, rows: list) -> dict:
    """將二維陣列寫入 Excel xlsx 工作表；需要 openpyxl。"""
    try:
        import openpyxl
        p = Path(path)
        p.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name or "Sheet1"
        for row in rows:
            ws.append(list(row) if isinstance(row, (list, tuple)) else [row])
        wb.save(p)
        return {"success": True, "path": str(p.resolve()), "sheet": ws.title, "rows": len(rows)}
    except Exception as e:
        return {"error": str(e)}


# =====================================================================
# GOOGLE 實體服務對接層（Gmail, Calendar, Drive, Sheets）
# =====================================================================

@tool()
def gmail_send(to: str, subject: str, body: str) -> dict:
    """發送電子郵件"""
    try:
        service = get_google_service('gmail', 'v1')
        import base64
        from email.mime.text import MIMEText
        message = MIMEText(body)
        message['to'] = to
        message['subject'] = subject
        raw = base64.urlsafe_b64encode(message.as_bytes()).decode()
        res = service.users().messages().send(userId='me', body={'raw': raw}).execute()
        return {"success": True, "id": res.get("id")}
    except Exception as e:
        return {"error": str(e)}

@tool()
def gmail_read(query: str) -> dict:
    """讀取郵件（支援搜尋條件，如 'from:someone'）"""
    try:
        service = get_google_service('gmail', 'v1')
        res = service.users().messages().list(userId='me', q=query, maxResults=1).execute()
        messages = res.get('messages', [])
        if not messages: return {"found": False, "message": "找不到符合的郵件", "query": query}
        msg_info = service.users().messages().get(userId='me', id=messages[0]['id']).execute()
        headers = {
            h.get("name", "").lower(): h.get("value", "")
            for h in msg_info.get("payload", {}).get("headers", [])
        }
        return {
            "found": True,
            "id": msg_info.get("id"),
            "thread_id": msg_info.get("threadId"),
            "from": headers.get("from", ""),
            "to": headers.get("to", ""),
            "subject": headers.get("subject", ""),
            "date": headers.get("date", ""),
            "snippet": msg_info.get("snippet", ""),
            "query": query,
        }
    except Exception as e:
        return {"error": str(e)}

@tool()
def gmail_list(max_results: int = 5) -> dict:
    """列出收件匣最新郵件"""
    try:
        service = get_google_service('gmail', 'v1')
        res = service.users().messages().list(userId='me', maxResults=max_results).execute()
        out = []
        for row in res.get('messages', []):
            msg = service.users().messages().get(
                userId='me',
                id=row['id'],
                format='metadata',
                metadataHeaders=['From', 'Subject', 'Date'],
            ).execute()
            headers = {
                h.get("name", "").lower(): h.get("value", "")
                for h in msg.get("payload", {}).get("headers", [])
            }
            out.append({
                "id": msg.get("id"),
                "thread_id": msg.get("threadId"),
                "from": headers.get("from", ""),
                "subject": headers.get("subject", ""),
                "date": headers.get("date", ""),
                "snippet": msg.get("snippet", ""),
            })
        return {"messages": out, "count": len(out)}
    except Exception as e:
        return {"error": str(e)}

@tool()
def gmail_delete(message_id: str) -> dict:
    """將郵件移至垃圾桶（刪除）"""
    try:
        service = get_google_service('gmail', 'v1')
        service.users().messages().trash(userId='me', id=message_id).execute()
        return {"success": True}
    except Exception as e:
        return {"error": str(e)}

@tool()
def gmail_reply(message_id: str, body: str) -> dict:
    """回覆指定 ID 的郵件"""
    try:
        service = get_google_service('gmail', 'v1')
        origin = service.users().messages().get(userId='me', id=message_id).execute()
        headers = origin.get('payload', {}).get('headers', [])
        subject = next((h['value'] for h in headers if h['name'].lower() == 'subject'), 'Re:')
        to = next((h['value'] for h in headers if h['name'].lower() == 'from'), '')

        if not subject.startswith('Re:'): subject = 'Re: ' + subject

        import base64
        from email.mime.text import MIMEText
        message = MIMEText(body)
        message['to'] = to
        message['subject'] = subject
        raw = base64.urlsafe_b64encode(message.as_bytes()).decode()
        res = service.users().messages().send(userId='me', body={'raw': raw, 'threadId': origin.get('threadId')}).execute()
        return {"success": True, "id": res.get("id")}
    except Exception as e:
        return {"error": str(e)}

@tool()
def calendar_list_events(date_start: str, date_end: str) -> dict:
    """列出指定日期範圍的行程 (ISO 格式字串，例如 '2026-05-01T00:00:00Z')"""
    try:
        service = get_google_service('calendar', 'v3')
        events_result = service.events().list(calendarId='primary', timeMin=date_start, timeMax=date_end, singleEvents=True, orderBy='startTime').execute()
        return {"events": events_result.get('items', [])}
    except Exception as e:
        return {"error": str(e)}

@tool()
def calendar_create_event(summary: str, start_time: str, end_time: str) -> dict:
    """建立新日曆行程 (時間請傳入 ISO 格式，如 '2026-05-25T10:00:00+08:00')"""
    try:
        service = get_google_service('calendar', 'v3')
        event = {
            'summary': summary,
            'start': {'dateTime': start_time, 'timeZone': 'Asia/Taipei'},
            'end': {'dateTime': end_time, 'timeZone': 'Asia/Taipei'},
        }
        event = service.events().insert(calendarId='primary', body=event).execute()
        return {"success": True, "htmlLink": event.get('htmlLink'), "id": event.get('id')}
    except Exception as e:
        return {"error": str(e)}

@tool()
def calendar_update_event(event_id: str, summary: str) -> dict:
    """更新現有行程的標題"""
    try:
        service = get_google_service('calendar', 'v3')
        event = service.events().get(calendarId='primary', eventId=event_id).execute()
        event['summary'] = summary
        updated_event = service.events().update(calendarId='primary', eventId=event_id, body=event).execute()
        return {"success": True, "updatedTime": updated_event.get('updated')}
    except Exception as e:
        return {"error": str(e)}

@tool()
def calendar_delete_event(event_id: str) -> dict:
    """刪除行程"""
    try:
        service = get_google_service('calendar', 'v3')
        service.events().delete(calendarId='primary', eventId=event_id).execute()
        return {"success": True}
    except Exception as e:
        return {"error": str(e)}

@tool()
def drive_list_files(query: str = "") -> dict:
    """列出 Drive 檔案（支援 MimeType 或名稱篩選）"""
    try:
        service = get_google_service('drive', 'v3')
        res = service.files().list(q=query, fields="nextPageToken, files(id, name, mimeType)").execute()
        return {"files": res.get('files', [])}
    except Exception as e:
        return {"error": str(e)}

@tool()
def drive_read_file(file_id: str) -> dict:
    """讀取雲端硬碟檔案內容（純文字或匯出 Google Docs 為純文字）"""
    try:
        service = get_google_service('drive', 'v3')
        meta = service.files().get(fileId=file_id).execute()
        mime = meta.get('mimeType', '')

        if 'application/vnd.google-apps' in mime:
            res = service.files().export_media(fileId=file_id, mimeType='text/plain').execute()
        else:
            res = service.files().get_media(fileId=file_id).execute()

        return {"name": meta.get("name"), "content": res.decode('utf-8', errors='ignore')}
    except Exception as e:
        return {"error": str(e)}

@tool()
def drive_upload_file(local_path: str, google_drive_folder_id: str = "") -> dict:
    """上傳本地檔案至 Drive"""
    try:
        service = get_google_service('drive', 'v3')
        p = Path(local_path)
        if not p.exists(): return {"error": f"找不到本地檔案 {local_path}"}

        meta = {'name': p.name}
        if google_drive_folder_id:
            meta['parents'] = [google_drive_folder_id]

        from googleapiclient.http import MediaFileUpload
        media = MediaFileUpload(local_path, resumable=True)
        file = service.files().create(body=meta, media_body=media, fields='id').execute()
        return {"success": True, "file_id": file.get('id')}
    except Exception as e:
        return {"error": str(e)}

@tool()
def drive_create_doc(title: str, content: str) -> dict:
    """建立新 Google Doc 文件"""
    try:
        drive_service = get_google_service('drive', 'v3')
        meta = {'name': title, 'mimeType': 'application/vnd.google-apps.document'}
        file = drive_service.files().create(body=meta).execute()
        doc_id = file.get('id')

        docs_service = get_google_service('docs', 'v1')
        requests = [{'insertText': {'location': {'index': 1}, 'text': content}}]
        docs_service.documents().batchUpdate(documentId=doc_id, body={'requests': requests}).execute()
        return {"success": True, "doc_id": doc_id}
    except Exception as e:
        return {"error": str(e)}

@tool()
def drive_share_file(file_id: str, email: str, role: str = "reader") -> dict:
    """設定雲端硬碟檔案分享權限 (role: reader, writer)"""
    try:
        service = get_google_service('drive', 'v3')
        user_permission = {'type': 'user', 'role': role, 'emailAddress': email}
        service.permissions().create(fileId=file_id, body=user_permission, fields='id').execute()
        return {"success": True}
    except Exception as e:
        return {"error": str(e)}

@tool()
def sheets_read_range(spreadsheet_id: str, range_name: str) -> dict:
    """讀取指定 Google 工作表範圍的資料 (如 'Sheet1!A1:D10')"""
    try:
        service = get_google_service('sheets', 'v4')
        res = service.spreadsheets().values().get(spreadsheetId=spreadsheet_id, range=range_name).execute()
        return {"values": res.get('values', [])}
    except Exception as e:
        return {"error": str(e)}

@tool()
def sheets_write_range(spreadsheet_id: str, range_name: str, values: list) -> dict:
    """覆寫資料至指定 Google 工作表範圍"""
    try:
        service = get_google_service('sheets', 'v4')
        body = {'values': values}
        res = service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id, range=range_name,
            valueInputOption='USER_ENTERED', body=body).execute()
        return {"success": True, "updatedCells": res.get('updatedCells')}
    except Exception as e:
        return {"error": str(e)}

@tool()
def sheets_append_row(spreadsheet_id: str, sheet_name: str, values: list) -> dict:
    """在工作表末尾新增一列資料"""
    try:
        service = get_google_service('sheets', 'v4')
        body = {'values': [values]}
        res = service.spreadsheets().values().append(
            spreadsheetId=spreadsheet_id, range=f"{sheet_name}!A1",
            valueInputOption='USER_ENTERED', insertDataOption='INSERT_ROWS', body=body).execute()
        return {"success": True, "updates": res.get('updates')}
    except Exception as e:
        return {"error": str(e)}

@tool()
def sheets_create(title: str) -> dict:
    """建立全新 Google 試算表"""
    try:
        service = get_google_service('sheets', 'v4')
        spreadsheet = {'properties': {'title': title}}
        res = service.spreadsheets().create(body=spreadsheet, fields='spreadsheetId').execute()
        return {"success": True, "spreadsheetId": res.get('spreadsheetId')}
    except Exception as e:
        return {"error": str(e)}


# =====================================================================
# 記憶工具（註冊進 registry，讓額外模型也能呼叫）
# =====================================================================

_mem_store = None


def close_tool_memory_store() -> None:
    global _mem_store
    store = _mem_store
    _mem_store = None
    if store and hasattr(store, "close"):
        try:
            store.close()
        except Exception:
            pass


atexit.register(close_tool_memory_store)

def _get_mem_store():
    global _mem_store
    if _mem_store is None:
        try:
            from tools.memory.store import MemoryStore
            _mem_store = MemoryStore(
                host=os.environ.get("QDRANT_HOST", "127.0.0.1"),
                port=int(os.environ.get("QDRANT_PORT", "6333")),
            )
        except Exception:
            return None
    return _mem_store


@tool(schema_override={
    "description": "將重要資訊寫入長期記憶",
    "properties": {
        "category": {"enum": ["personal","preference","technical","project","event","credential","other"]},
    },
})
def store_memory(memory: str, category: str = "other", importance: int = 3) -> dict:
    """將對話中出現的重要資訊寫入長期記憶。"""
    store = _get_mem_store()
    if not store:
        return {"error": "記憶系統離線（Qdrant 未啟動）"}
    user_id = os.environ.get("RAPHAEL_USER", "wayne")
    mid = store.store(memory, user_id=user_id, category=category, importance=importance)
    return {"stored": True, "id": mid[:8]}


@tool()
def recall_memory(query: str, limit: int = 5) -> dict:
    """語意搜尋長期記憶。"""
    store = _get_mem_store()
    if not store:
        return {"error": "記憶系統離線（Qdrant 未啟動）"}
    user_id = os.environ.get("RAPHAEL_USER", "wayne")
    results = store.search(query, user_id=user_id, limit=limit)
    return {"results": results, "count": len(results)}
